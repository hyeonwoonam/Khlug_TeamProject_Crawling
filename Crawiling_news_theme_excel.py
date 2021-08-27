from openpyxl.workbook.workbook import Workbook
import requests
from bs4 import BeautifulSoup
import time
from openpyxl import workbook

date = input("Crawling을 진행할 날짜를 입력해주세요. (YYYYMMDD) : ")

wb = Workbook()
time = time.localtime()

headers = {"User-Agent":"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36 Edg/91.0.864.64"}

def create_soup(url) :
    res = requests.get(url, headers=headers)
    res.raise_for_status()
    soup = BeautifulSoup(res.text, "lxml")
    return soup

def create_paging_soup(paging_url) :
    res = requests.get(paging_url, headers=headers)
    res.raise_for_status()
    paging_soup = BeautifulSoup(res.text, "lxml")
    return paging_soup

def theme_100_264() :   # 정치 > 청와대
    sid1_code = str(100)
    sid2_code = str(264)

    ws = wb.active
    ws.title = "정치_청와대"
    ws.append(["index", "제목", "링크", "언론사"])

    def scrape() :
        page_index_1 = 0
        page_index_2 = 10

        for i in range(int(page_start), int(page_end)+1) :
            url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(i)

            soup = create_soup(url)
            news_list1 = soup.find("ul", attrs={"class":"type06_headline"}).find_all("li")
            for index, news1 in enumerate(news_list1) :
                a_img = 0
                img = news1.find("img")
                if img :
                    a_img = 1

                a_tag1 = news1.find_all("a")[a_img]
                title = a_tag1.get_text().strip()
                link = a_tag1["href"]
                press = news1.find("span", attrs={"class":"writing"}).get_text()

                page_index_1 = page_index_1 + 1

                ws.append([page_index_1, title, link, press])

            page_index_1 = page_index_1 + index + 1

            if(soup.find("ul", attrs={"class":"type06"})) :
                news_list2 = soup.find("ul", attrs={"class":"type06"}).find_all("li")
                for index, news2 in enumerate(news_list2) :
                        a_img = 0
                        img = news2.find("img")
                        if img :
                            a_img = 1

                        a_tag2 = news2.find_all("a")[a_img]
                        title = a_tag2.get_text().strip()
                        link = a_tag2["href"]
                        press = news2.find("span", attrs={"class":"writing"}).get_text()

                        page_index_2 = page_index_2 + 1

                        ws.append([page_index_2, title, link, press])

                page_index_2 = page_index_2 + index + 1

    page_number = 1
    paging_url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(page_number)
    paging_soup = create_paging_soup(paging_url)
    page_start = paging_soup.find("div", attrs={"class":"paging"}).find("strong").get_text()

    if(paging_soup.find("div", attrs={"class":"paging"}).find("a")) :
        page = paging_soup.find("div", attrs={"class":"paging"}).find_all("a")[-1]
        page_end = page.get_text()
    else :
        page_end = page_start

    if(page_end == "다음") :
        while(page_end == "다음") :
            page_number = page_number + 10
            paging_url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(page_number)
            paging_soup = create_paging_soup(paging_url)
            page = paging_soup.find("div", attrs={"class":"paging"}).find_all("a")[-1]
            page_end = page.get_text()

            if(page_end != "다음") :
                if(page_end == "이전") :
                    page_end = paging_soup.find("div", attrs={"class":"paging"}).find("strong").get_text()
                    scrape()
                else :
                    scrape()
    elif(page_end != "다음") :
        scrape()

def theme_100_265() :   # 정치 > 국회/정당
    sid1_code = str(100)
    sid2_code = str(265)

    ws1 = wb.create_sheet("정치_국회,정당")
    ws1 = wb["정치_국회,정당"]
    ws1.append(["index", "제목", "링크", "언론사"])

    def scrape() :
        page_index_1 = 0
        page_index_2 = 10

        for i in range(int(page_start), int(page_end)+1) :
            url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(i)

            soup = create_soup(url)
            news_list1 = soup.find("ul", attrs={"class":"type06_headline"}).find_all("li")
            for index, news1 in enumerate(news_list1) :
                a_img = 0
                img = news1.find("img")
                if img :
                    a_img = 1

                a_tag1 = news1.find_all("a")[a_img]
                title = a_tag1.get_text().strip()
                link = a_tag1["href"]
                press = news1.find("span", attrs={"class":"writing"}).get_text()

                page_index_1 = page_index_1 + 1

                ws1.append([page_index_1, title, link, press])

            page_index_1 = page_index_1 + index + 1

            if(soup.find("ul", attrs={"class":"type06"})) :
                news_list2 = soup.find("ul", attrs={"class":"type06"}).find_all("li")
                for index, news2 in enumerate(news_list2) :
                        a_img = 0
                        img = news2.find("img")
                        if img :
                            a_img = 1

                        a_tag2 = news2.find_all("a")[a_img]
                        title = a_tag2.get_text().strip()
                        link = a_tag2["href"]
                        press = news2.find("span", attrs={"class":"writing"}).get_text()

                        page_index_2 = page_index_2 + 1

                        ws1.append([page_index_2, title, link, press])

                page_index_2 = page_index_2 + index + 1

    page_number = 1
    paging_url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(page_number)
    paging_soup = create_paging_soup(paging_url)
    page_start = paging_soup.find("div", attrs={"class":"paging"}).find("strong").get_text()

    if(paging_soup.find("div", attrs={"class":"paging"}).find("a")) :
        page = paging_soup.find("div", attrs={"class":"paging"}).find_all("a")[-1]
        page_end = page.get_text()
    else :
        page_end = page_start

    if(page_end == "다음") :
        while(page_end == "다음") :
            page_number = page_number + 10
            paging_url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(page_number)
            paging_soup = create_paging_soup(paging_url)
            page = paging_soup.find("div", attrs={"class":"paging"}).find_all("a")[-1]
            page_end = page.get_text()

            if(page_end != "다음") :
                if(page_end == "이전") :
                    page_end = paging_soup.find("div", attrs={"class":"paging"}).find("strong").get_text()
                    scrape()
                else :
                    scrape()
    elif(page_end != "다음") :
        scrape()

def theme_100_268() :   # 정치 > 북한
    sid1_code = str(100)
    sid2_code = str(268)

    ws1 = wb.create_sheet("정치_북한")
    ws1 = wb["정치_북한"]
    ws1.append(["index", "제목", "링크", "언론사"])

    def scrape() :
        page_index_1 = 0
        page_index_2 = 10

        for i in range(int(page_start), int(page_end)+1) :
            url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(i)

            soup = create_soup(url)
            news_list1 = soup.find("ul", attrs={"class":"type06_headline"}).find_all("li")
            for index, news1 in enumerate(news_list1) :
                a_img = 0
                img = news1.find("img")
                if img :
                    a_img = 1

                a_tag1 = news1.find_all("a")[a_img]
                title = a_tag1.get_text().strip()
                link = a_tag1["href"]
                press = news1.find("span", attrs={"class":"writing"}).get_text()

                page_index_1 = page_index_1 + 1

                ws1.append([page_index_1, title, link, press])

            page_index_1 = page_index_1 + index + 1

            if(soup.find("ul", attrs={"class":"type06"})) :
                news_list2 = soup.find("ul", attrs={"class":"type06"}).find_all("li")
                for index, news2 in enumerate(news_list2) :
                        a_img = 0
                        img = news2.find("img")
                        if img :
                            a_img = 1

                        a_tag2 = news2.find_all("a")[a_img]
                        title = a_tag2.get_text().strip()
                        link = a_tag2["href"]
                        press = news2.find("span", attrs={"class":"writing"}).get_text()

                        page_index_2 = page_index_2 + 1

                        ws1.append([page_index_2, title, link, press])

                page_index_2 = page_index_2 + index + 1

    page_number = 1
    paging_url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(page_number)
    paging_soup = create_paging_soup(paging_url)
    page_start = paging_soup.find("div", attrs={"class":"paging"}).find("strong").get_text()

    if(paging_soup.find("div", attrs={"class":"paging"}).find("a")) :
        page = paging_soup.find("div", attrs={"class":"paging"}).find_all("a")[-1]
        page_end = page.get_text()
    else :
        page_end = page_start

    if(page_end == "다음") :
        while(page_end == "다음") :
            page_number = page_number + 10
            paging_url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(page_number)
            paging_soup = create_paging_soup(paging_url)
            page = paging_soup.find("div", attrs={"class":"paging"}).find_all("a")[-1]
            page_end = page.get_text()

            if(page_end != "다음") :
                if(page_end == "이전") :
                    page_end = paging_soup.find("div", attrs={"class":"paging"}).find("strong").get_text()
                    scrape()
                else :
                    scrape()
    elif(page_end != "다음") :
        scrape()

def theme_100_266() :   # 정치 > 행정
    sid1_code = str(100)
    sid2_code = str(266)

    ws1 = wb.create_sheet("정치_행정")
    ws1 = wb["정치_행정"]
    ws1.append(["index", "제목", "링크", "언론사"])

    def scrape() :
        page_index_1 = 0
        page_index_2 = 10

        for i in range(int(page_start), int(page_end)+1) :
            url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(i)

            soup = create_soup(url)
            news_list1 = soup.find("ul", attrs={"class":"type06_headline"}).find_all("li")
            for index, news1 in enumerate(news_list1) :
                a_img = 0
                img = news1.find("img")
                if img :
                    a_img = 1

                a_tag1 = news1.find_all("a")[a_img]
                title = a_tag1.get_text().strip()
                link = a_tag1["href"]
                press = news1.find("span", attrs={"class":"writing"}).get_text()

                page_index_1 = page_index_1 + 1

                ws1.append([page_index_1, title, link, press])

            page_index_1 = page_index_1 + index + 1

            if(soup.find("ul", attrs={"class":"type06"})) :
                news_list2 = soup.find("ul", attrs={"class":"type06"}).find_all("li")
                for index, news2 in enumerate(news_list2) :
                        a_img = 0
                        img = news2.find("img")
                        if img :
                            a_img = 1

                        a_tag2 = news2.find_all("a")[a_img]
                        title = a_tag2.get_text().strip()
                        link = a_tag2["href"]
                        press = news2.find("span", attrs={"class":"writing"}).get_text()

                        page_index_2 = page_index_2 + 1

                        ws1.append([page_index_2, title, link, press])

                page_index_2 = page_index_2 + index + 1

    page_number = 1
    paging_url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(page_number)
    paging_soup = create_paging_soup(paging_url)
    page_start = paging_soup.find("div", attrs={"class":"paging"}).find("strong").get_text()

    if(paging_soup.find("div", attrs={"class":"paging"}).find("a")) :
        page = paging_soup.find("div", attrs={"class":"paging"}).find_all("a")[-1]
        page_end = page.get_text()
    else :
        page_end = page_start

    if(page_end == "다음") :
        while(page_end == "다음") :
            page_number = page_number + 10
            paging_url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(page_number)
            paging_soup = create_paging_soup(paging_url)
            page = paging_soup.find("div", attrs={"class":"paging"}).find_all("a")[-1]
            page_end = page.get_text()

            if(page_end != "다음") :
                if(page_end == "이전") :
                    page_end = paging_soup.find("div", attrs={"class":"paging"}).find("strong").get_text()
                    scrape()
                else :
                    scrape()
    elif(page_end != "다음") :
        scrape()

def theme_100_267() :   # 정치 > 국방/외교
    sid1_code = str(100)
    sid2_code = str(267)

    ws1 = wb.create_sheet("정치_국방,외교")
    ws1 = wb["정치_국방,외교"]
    ws1.append(["index", "제목", "링크", "언론사"])

    def scrape() :
        page_index_1 = 0
        page_index_2 = 10

        for i in range(int(page_start), int(page_end)+1) :
            url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(i)

            soup = create_soup(url)
            news_list1 = soup.find("ul", attrs={"class":"type06_headline"}).find_all("li")
            for index, news1 in enumerate(news_list1) :
                a_img = 0
                img = news1.find("img")
                if img :
                    a_img = 1

                a_tag1 = news1.find_all("a")[a_img]
                title = a_tag1.get_text().strip()
                link = a_tag1["href"]
                press = news1.find("span", attrs={"class":"writing"}).get_text()

                page_index_1 = page_index_1 + 1

                ws1.append([page_index_1, title, link, press])

            page_index_1 = page_index_1 + index + 1

            if(soup.find("ul", attrs={"class":"type06"})) :
                news_list2 = soup.find("ul", attrs={"class":"type06"}).find_all("li")
                for index, news2 in enumerate(news_list2) :
                        a_img = 0
                        img = news2.find("img")
                        if img :
                            a_img = 1

                        a_tag2 = news2.find_all("a")[a_img]
                        title = a_tag2.get_text().strip()
                        link = a_tag2["href"]
                        press = news2.find("span", attrs={"class":"writing"}).get_text()

                        page_index_2 = page_index_2 + 1

                        ws1.append([page_index_2, title, link, press])

                page_index_2 = page_index_2 + index + 1

    page_number = 1
    paging_url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(page_number)
    paging_soup = create_paging_soup(paging_url)
    page_start = paging_soup.find("div", attrs={"class":"paging"}).find("strong").get_text()

    if(paging_soup.find("div", attrs={"class":"paging"}).find("a")) :
        page = paging_soup.find("div", attrs={"class":"paging"}).find_all("a")[-1]
        page_end = page.get_text()
    else :
        page_end = page_start

    if(page_end == "다음") :
        while(page_end == "다음") :
            page_number = page_number + 10
            paging_url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(page_number)
            paging_soup = create_paging_soup(paging_url)
            page = paging_soup.find("div", attrs={"class":"paging"}).find_all("a")[-1]
            page_end = page.get_text()

            if(page_end != "다음") :
                if(page_end == "이전") :
                    page_end = paging_soup.find("div", attrs={"class":"paging"}).find("strong").get_text()
                    scrape()
                else :
                    scrape()
    elif(page_end != "다음") :
        scrape()

def theme_100_269() :   # 정치 > 정치일반
    sid1_code = str(100)
    sid2_code = str(269)

    ws1 = wb.create_sheet("정치_정치일반")
    ws1 = wb["정치_정치일반"]
    ws1.append(["index", "제목", "링크", "언론사"])

    def scrape() :
        page_index_1 = 0
        page_index_2 = 10

        for i in range(int(page_start), int(page_end)+1) :
            url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(i)

            soup = create_soup(url)
            news_list1 = soup.find("ul", attrs={"class":"type06_headline"}).find_all("li")
            for index, news1 in enumerate(news_list1) :
                a_img = 0
                img = news1.find("img")
                if img :
                    a_img = 1

                a_tag1 = news1.find_all("a")[a_img]
                title = a_tag1.get_text().strip()
                link = a_tag1["href"]
                press = news1.find("span", attrs={"class":"writing"}).get_text()

                page_index_1 = page_index_1 + 1

                ws1.append([page_index_1, title, link, press])

            page_index_1 = page_index_1 + index + 1

            if(soup.find("ul", attrs={"class":"type06"})) :
                news_list2 = soup.find("ul", attrs={"class":"type06"}).find_all("li")
                for index, news2 in enumerate(news_list2) :
                        a_img = 0
                        img = news2.find("img")
                        if img :
                            a_img = 1

                        a_tag2 = news2.find_all("a")[a_img]
                        title = a_tag2.get_text().strip()
                        link = a_tag2["href"]
                        press = news2.find("span", attrs={"class":"writing"}).get_text()

                        page_index_2 = page_index_2 + 1

                        ws1.append([page_index_2, title, link, press])

                page_index_2 = page_index_2 + index + 1

    page_number = 1
    paging_url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(page_number)
    paging_soup = create_paging_soup(paging_url)
    page_start = paging_soup.find("div", attrs={"class":"paging"}).find("strong").get_text()

    if(paging_soup.find("div", attrs={"class":"paging"}).find("a")) :
        page = paging_soup.find("div", attrs={"class":"paging"}).find_all("a")[-1]
        page_end = page.get_text()
    else :
        page_end = page_start

    if(page_end == "다음") :
        while(page_end == "다음") :
            page_number = page_number + 10
            paging_url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(page_number)
            paging_soup = create_paging_soup(paging_url)
            page = paging_soup.find("div", attrs={"class":"paging"}).find_all("a")[-1]
            page_end = page.get_text()

            if(page_end != "다음") :
                if(page_end == "이전") :
                    page_end = paging_soup.find("div", attrs={"class":"paging"}).find("strong").get_text()
                    scrape()
                else :
                    scrape()
    elif(page_end != "다음") :
        scrape()

def theme_101_259() :   # 경제 > 금융
    sid1_code = str(101)
    sid2_code = str(259)

    ws1 = wb.create_sheet("경제_금융")
    ws1 = wb["경제_금융"]
    ws1.append(["index", "제목", "링크", "언론사"])

    def scrape() :
        page_index_1 = 0
        page_index_2 = 10

        for i in range(int(page_start), int(page_end)+1) :
            url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(i)

            soup = create_soup(url)
            news_list1 = soup.find("ul", attrs={"class":"type06_headline"}).find_all("li")
            for index, news1 in enumerate(news_list1) :
                a_img = 0
                img = news1.find("img")
                if img :
                    a_img = 1

                a_tag1 = news1.find_all("a")[a_img]
                title = a_tag1.get_text().strip()
                link = a_tag1["href"]
                press = news1.find("span", attrs={"class":"writing"}).get_text()

                page_index_1 = page_index_1 + 1

                ws1.append([page_index_1, title, link, press])

            page_index_1 = page_index_1 + index + 1

            if(soup.find("ul", attrs={"class":"type06"})) :
                news_list2 = soup.find("ul", attrs={"class":"type06"}).find_all("li")
                for index, news2 in enumerate(news_list2) :
                        a_img = 0
                        img = news2.find("img")
                        if img :
                            a_img = 1

                        a_tag2 = news2.find_all("a")[a_img]
                        title = a_tag2.get_text().strip()
                        link = a_tag2["href"]
                        press = news2.find("span", attrs={"class":"writing"}).get_text()

                        page_index_2 = page_index_2 + 1

                        ws1.append([page_index_2, title, link, press])

                page_index_2 = page_index_2 + index + 1

    page_number = 1
    paging_url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(page_number)
    paging_soup = create_paging_soup(paging_url)
    page_start = paging_soup.find("div", attrs={"class":"paging"}).find("strong").get_text()

    if(paging_soup.find("div", attrs={"class":"paging"}).find("a")) :
        page = paging_soup.find("div", attrs={"class":"paging"}).find_all("a")[-1]
        page_end = page.get_text()
    else :
        page_end = page_start

    if(page_end == "다음") :
        while(page_end == "다음") :
            page_number = page_number + 10
            paging_url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(page_number)
            paging_soup = create_paging_soup(paging_url)
            page = paging_soup.find("div", attrs={"class":"paging"}).find_all("a")[-1]
            page_end = page.get_text()

            if(page_end != "다음") :
                if(page_end == "이전") :
                    page_end = paging_soup.find("div", attrs={"class":"paging"}).find("strong").get_text()
                    scrape()
                else :
                    scrape()
    elif(page_end != "다음") :
        scrape()

def theme_101_258() :   # 경제 > 증권
    sid1_code = str(101)
    sid2_code = str(258)

    ws1 = wb.create_sheet("경제_증권")
    ws1 = wb["경제_증권"]
    ws1.append(["index", "제목", "링크", "언론사"])

    def scrape() :
        page_index_1 = 0
        page_index_2 = 10

        for i in range(int(page_start), int(page_end)+1) :
            url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(i)

            soup = create_soup(url)
            news_list1 = soup.find("ul", attrs={"class":"type06_headline"}).find_all("li")
            for index, news1 in enumerate(news_list1) :
                a_img = 0
                img = news1.find("img")
                if img :
                    a_img = 1

                a_tag1 = news1.find_all("a")[a_img]
                title = a_tag1.get_text().strip()
                link = a_tag1["href"]
                press = news1.find("span", attrs={"class":"writing"}).get_text()

                page_index_1 = page_index_1 + 1

                ws1.append([page_index_1, title, link, press])

            page_index_1 = page_index_1 + index + 1

            if(soup.find("ul", attrs={"class":"type06"})) :
                news_list2 = soup.find("ul", attrs={"class":"type06"}).find_all("li")
                for index, news2 in enumerate(news_list2) :
                        a_img = 0
                        img = news2.find("img")
                        if img :
                            a_img = 1

                        a_tag2 = news2.find_all("a")[a_img]
                        title = a_tag2.get_text().strip()
                        link = a_tag2["href"]
                        press = news2.find("span", attrs={"class":"writing"}).get_text()

                        page_index_2 = page_index_2 + 1

                        ws1.append([page_index_2, title, link, press])

                page_index_2 = page_index_2 + index + 1

    page_number = 1
    paging_url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(page_number)
    paging_soup = create_paging_soup(paging_url)
    page_start = paging_soup.find("div", attrs={"class":"paging"}).find("strong").get_text()

    if(paging_soup.find("div", attrs={"class":"paging"}).find("a")) :
        page = paging_soup.find("div", attrs={"class":"paging"}).find_all("a")[-1]
        page_end = page.get_text()
    else :
        page_end = page_start

    if(page_end == "다음") :
        while(page_end == "다음") :
            page_number = page_number + 10
            paging_url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(page_number)
            paging_soup = create_paging_soup(paging_url)
            page = paging_soup.find("div", attrs={"class":"paging"}).find_all("a")[-1]
            page_end = page.get_text()

            if(page_end != "다음") :
                if(page_end == "이전") :
                    page_end = paging_soup.find("div", attrs={"class":"paging"}).find("strong").get_text()
                    scrape()
                else :
                    scrape()
    elif(page_end != "다음") :
        scrape()

def theme_101_261() :   # 경제 > 산업/재계
    sid1_code = str(101)
    sid2_code = str(261)

    ws1 = wb.create_sheet("경제_산업,재계")
    ws1 = wb["경제_산업,재계"]
    ws1.append(["index", "제목", "링크", "언론사"])

    def scrape() :
        page_index_1 = 0
        page_index_2 = 10

        for i in range(int(page_start), int(page_end)+1) :
            url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(i)

            soup = create_soup(url)
            news_list1 = soup.find("ul", attrs={"class":"type06_headline"}).find_all("li")
            for index, news1 in enumerate(news_list1) :
                a_img = 0
                img = news1.find("img")
                if img :
                    a_img = 1

                a_tag1 = news1.find_all("a")[a_img]
                title = a_tag1.get_text().strip()
                link = a_tag1["href"]
                press = news1.find("span", attrs={"class":"writing"}).get_text()

                page_index_1 = page_index_1 + 1

                ws1.append([page_index_1, title, link, press])

            page_index_1 = page_index_1 + index + 1

            if(soup.find("ul", attrs={"class":"type06"})) :
                news_list2 = soup.find("ul", attrs={"class":"type06"}).find_all("li")
                for index, news2 in enumerate(news_list2) :
                        a_img = 0
                        img = news2.find("img")
                        if img :
                            a_img = 1

                        a_tag2 = news2.find_all("a")[a_img]
                        title = a_tag2.get_text().strip()
                        link = a_tag2["href"]
                        press = news2.find("span", attrs={"class":"writing"}).get_text()

                        page_index_2 = page_index_2 + 1

                        ws1.append([page_index_2, title, link, press])

                page_index_2 = page_index_2 + index + 1

    page_number = 1
    paging_url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(page_number)
    paging_soup = create_paging_soup(paging_url)
    page_start = paging_soup.find("div", attrs={"class":"paging"}).find("strong").get_text()

    if(paging_soup.find("div", attrs={"class":"paging"}).find("a")) :
        page = paging_soup.find("div", attrs={"class":"paging"}).find_all("a")[-1]
        page_end = page.get_text()
    else :
        page_end = page_start

    if(page_end == "다음") :
        while(page_end == "다음") :
            page_number = page_number + 10
            paging_url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(page_number)
            paging_soup = create_paging_soup(paging_url)
            page = paging_soup.find("div", attrs={"class":"paging"}).find_all("a")[-1]
            page_end = page.get_text()

            if(page_end != "다음") :
                if(page_end == "이전") :
                    page_end = paging_soup.find("div", attrs={"class":"paging"}).find("strong").get_text()
                    scrape()
                else :
                    scrape()
    elif(page_end != "다음") :
        scrape()

def theme_101_771() :   # 경제 > 중기/벤처
    sid1_code = str(101)
    sid2_code = str(771)

    ws1 = wb.create_sheet("경제_중기,벤처")
    ws1 = wb["경제_중기,벤처"]
    ws1.append(["index", "제목", "링크", "언론사"])

    def scrape() :
        page_index_1 = 0
        page_index_2 = 10

        for i in range(int(page_start), int(page_end)+1) :
            url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(i)

            soup = create_soup(url)
            news_list1 = soup.find("ul", attrs={"class":"type06_headline"}).find_all("li")
            for index, news1 in enumerate(news_list1) :
                a_img = 0
                img = news1.find("img")
                if img :
                    a_img = 1

                a_tag1 = news1.find_all("a")[a_img]
                title = a_tag1.get_text().strip()
                link = a_tag1["href"]
                press = news1.find("span", attrs={"class":"writing"}).get_text()

                page_index_1 = page_index_1 + 1

                ws1.append([page_index_1, title, link, press])

            page_index_1 = page_index_1 + index + 1

            if(soup.find("ul", attrs={"class":"type06"})) :
                news_list2 = soup.find("ul", attrs={"class":"type06"}).find_all("li")
                for index, news2 in enumerate(news_list2) :
                        a_img = 0
                        img = news2.find("img")
                        if img :
                            a_img = 1

                        a_tag2 = news2.find_all("a")[a_img]
                        title = a_tag2.get_text().strip()
                        link = a_tag2["href"]
                        press = news2.find("span", attrs={"class":"writing"}).get_text()

                        page_index_2 = page_index_2 + 1

                        ws1.append([page_index_2, title, link, press])

                page_index_2 = page_index_2 + index + 1

    page_number = 1
    paging_url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(page_number)
    paging_soup = create_paging_soup(paging_url)
    page_start = paging_soup.find("div", attrs={"class":"paging"}).find("strong").get_text()

    if(paging_soup.find("div", attrs={"class":"paging"}).find("a")) :
        page = paging_soup.find("div", attrs={"class":"paging"}).find_all("a")[-1]
        page_end = page.get_text()
    else :
        page_end = page_start

    if(page_end == "다음") :
        while(page_end == "다음") :
            page_number = page_number + 10
            paging_url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(page_number)
            paging_soup = create_paging_soup(paging_url)
            page = paging_soup.find("div", attrs={"class":"paging"}).find_all("a")[-1]
            page_end = page.get_text()

            if(page_end != "다음") :
                if(page_end == "이전") :
                    page_end = paging_soup.find("div", attrs={"class":"paging"}).find("strong").get_text()
                    scrape()
                else :
                    scrape()
    elif(page_end != "다음") :
        scrape()

def theme_101_260() :   # 경제 > 부동산
    sid1_code = str(101)
    sid2_code = str(260)

    ws1 = wb.create_sheet("경제_부동산")
    ws1 = wb["경제_부동산"]
    ws1.append(["index", "제목", "링크", "언론사"])

    def scrape() :
        page_index_1 = 0
        page_index_2 = 10

        for i in range(int(page_start), int(page_end)+1) :
            url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(i)

            soup = create_soup(url)
            news_list1 = soup.find("ul", attrs={"class":"type06_headline"}).find_all("li")
            for index, news1 in enumerate(news_list1) :
                a_img = 0
                img = news1.find("img")
                if img :
                    a_img = 1

                a_tag1 = news1.find_all("a")[a_img]
                title = a_tag1.get_text().strip()
                link = a_tag1["href"]
                press = news1.find("span", attrs={"class":"writing"}).get_text()

                page_index_1 = page_index_1 + 1

                ws1.append([page_index_1, title, link, press])

            page_index_1 = page_index_1 + index + 1

            if(soup.find("ul", attrs={"class":"type06"})) :
                news_list2 = soup.find("ul", attrs={"class":"type06"}).find_all("li")
                for index, news2 in enumerate(news_list2) :
                        a_img = 0
                        img = news2.find("img")
                        if img :
                            a_img = 1

                        a_tag2 = news2.find_all("a")[a_img]
                        title = a_tag2.get_text().strip()
                        link = a_tag2["href"]
                        press = news2.find("span", attrs={"class":"writing"}).get_text()

                        page_index_2 = page_index_2 + 1

                        ws1.append([page_index_2, title, link, press])

                page_index_2 = page_index_2 + index + 1

    page_number = 1
    paging_url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(page_number)
    paging_soup = create_paging_soup(paging_url)
    page_start = paging_soup.find("div", attrs={"class":"paging"}).find("strong").get_text()

    if(paging_soup.find("div", attrs={"class":"paging"}).find("a")) :
        page = paging_soup.find("div", attrs={"class":"paging"}).find_all("a")[-1]
        page_end = page.get_text()
    else :
        page_end = page_start

    if(page_end == "다음") :
        while(page_end == "다음") :
            page_number = page_number + 10
            paging_url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(page_number)
            paging_soup = create_paging_soup(paging_url)
            page = paging_soup.find("div", attrs={"class":"paging"}).find_all("a")[-1]
            page_end = page.get_text()

            if(page_end != "다음") :
                if(page_end == "이전") :
                    page_end = paging_soup.find("div", attrs={"class":"paging"}).find("strong").get_text()
                    scrape()
                else :
                    scrape()
    elif(page_end != "다음") :
        scrape()

def theme_101_262() :   # 경제 > 글로벌경제
    sid1_code = str(101)
    sid2_code = str(262)

    ws1 = wb.create_sheet("경제_글로벌경제")
    ws1 = wb["경제_글로벌경제"]
    ws1.append(["index", "제목", "링크", "언론사"])

    def scrape() :
        page_index_1 = 0
        page_index_2 = 10

        for i in range(int(page_start), int(page_end)+1) :
            url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(i)

            soup = create_soup(url)
            news_list1 = soup.find("ul", attrs={"class":"type06_headline"}).find_all("li")
            for index, news1 in enumerate(news_list1) :
                a_img = 0
                img = news1.find("img")
                if img :
                    a_img = 1

                a_tag1 = news1.find_all("a")[a_img]
                title = a_tag1.get_text().strip()
                link = a_tag1["href"]
                press = news1.find("span", attrs={"class":"writing"}).get_text()

                page_index_1 = page_index_1 + 1

                ws1.append([page_index_1, title, link, press])

            page_index_1 = page_index_1 + index + 1

            if(soup.find("ul", attrs={"class":"type06"})) :
                news_list2 = soup.find("ul", attrs={"class":"type06"}).find_all("li")
                for index, news2 in enumerate(news_list2) :
                        a_img = 0
                        img = news2.find("img")
                        if img :
                            a_img = 1

                        a_tag2 = news2.find_all("a")[a_img]
                        title = a_tag2.get_text().strip()
                        link = a_tag2["href"]
                        press = news2.find("span", attrs={"class":"writing"}).get_text()

                        page_index_2 = page_index_2 + 1

                        ws1.append([page_index_2, title, link, press])

                page_index_2 = page_index_2 + index + 1

    page_number = 1
    paging_url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(page_number)
    paging_soup = create_paging_soup(paging_url)
    page_start = paging_soup.find("div", attrs={"class":"paging"}).find("strong").get_text()

    if(paging_soup.find("div", attrs={"class":"paging"}).find("a")) :
        page = paging_soup.find("div", attrs={"class":"paging"}).find_all("a")[-1]
        page_end = page.get_text()
    else :
        page_end = page_start

    if(page_end == "다음") :
        while(page_end == "다음") :
            page_number = page_number + 10
            paging_url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(page_number)
            paging_soup = create_paging_soup(paging_url)
            page = paging_soup.find("div", attrs={"class":"paging"}).find_all("a")[-1]
            page_end = page.get_text()

            if(page_end != "다음") :
                if(page_end == "이전") :
                    page_end = paging_soup.find("div", attrs={"class":"paging"}).find("strong").get_text()
                    scrape()
                else :
                    scrape()
    elif(page_end != "다음") :
        scrape()

def theme_101_310() :   # 경제 > 생활경제
    sid1_code = str(101)
    sid2_code = str(310)

    ws1 = wb.create_sheet("경제_생활경제")
    ws1 = wb["경제_생활경제"]
    ws1.append(["index", "제목", "링크", "언론사"])

    def scrape() :
        page_index_1 = 0
        page_index_2 = 10

        for i in range(int(page_start), int(page_end)+1) :
            url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(i)

            soup = create_soup(url)
            news_list1 = soup.find("ul", attrs={"class":"type06_headline"}).find_all("li")
            for index, news1 in enumerate(news_list1) :
                a_img = 0
                img = news1.find("img")
                if img :
                    a_img = 1

                a_tag1 = news1.find_all("a")[a_img]
                title = a_tag1.get_text().strip()
                link = a_tag1["href"]
                press = news1.find("span", attrs={"class":"writing"}).get_text()

                page_index_1 = page_index_1 + 1

                ws1.append([page_index_1, title, link, press])

            page_index_1 = page_index_1 + index + 1

            if(soup.find("ul", attrs={"class":"type06"})) :
                news_list2 = soup.find("ul", attrs={"class":"type06"}).find_all("li")
                for index, news2 in enumerate(news_list2) :
                        a_img = 0
                        img = news2.find("img")
                        if img :
                            a_img = 1

                        a_tag2 = news2.find_all("a")[a_img]
                        title = a_tag2.get_text().strip()
                        link = a_tag2["href"]
                        press = news2.find("span", attrs={"class":"writing"}).get_text()

                        page_index_2 = page_index_2 + 1

                        ws1.append([page_index_2, title, link, press])

                page_index_2 = page_index_2 + index + 1

    page_number = 1
    paging_url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(page_number)
    paging_soup = create_paging_soup(paging_url)
    page_start = paging_soup.find("div", attrs={"class":"paging"}).find("strong").get_text()

    if(paging_soup.find("div", attrs={"class":"paging"}).find("a")) :
        page = paging_soup.find("div", attrs={"class":"paging"}).find_all("a")[-1]
        page_end = page.get_text()
    else :
        page_end = page_start

    if(page_end == "다음") :
        while(page_end == "다음") :
            page_number = page_number + 10
            paging_url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(page_number)
            paging_soup = create_paging_soup(paging_url)
            page = paging_soup.find("div", attrs={"class":"paging"}).find_all("a")[-1]
            page_end = page.get_text()

            if(page_end != "다음") :
                if(page_end == "이전") :
                    page_end = paging_soup.find("div", attrs={"class":"paging"}).find("strong").get_text()
                    scrape()
                else :
                    scrape()
    elif(page_end != "다음") :
        scrape()

def theme_101_263() :   # 경제 > 경제일반
    sid1_code = str(101)
    sid2_code = str(263)

    ws1 = wb.create_sheet("경제_경제일반")
    ws1 = wb["경제_경제일반"]
    ws1.append(["index", "제목", "링크", "언론사"])

    def scrape() :
        page_index_1 = 0
        page_index_2 = 10

        for i in range(int(page_start), int(page_end)+1) :
            url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(i)

            soup = create_soup(url)
            news_list1 = soup.find("ul", attrs={"class":"type06_headline"}).find_all("li")
            for index, news1 in enumerate(news_list1) :
                a_img = 0
                img = news1.find("img")
                if img :
                    a_img = 1

                a_tag1 = news1.find_all("a")[a_img]
                title = a_tag1.get_text().strip()
                link = a_tag1["href"]
                press = news1.find("span", attrs={"class":"writing"}).get_text()

                page_index_1 = page_index_1 + 1

                ws1.append([page_index_1, title, link, press])

            page_index_1 = page_index_1 + index + 1

            if(soup.find("ul", attrs={"class":"type06"})) :
                news_list2 = soup.find("ul", attrs={"class":"type06"}).find_all("li")
                for index, news2 in enumerate(news_list2) :
                        a_img = 0
                        img = news2.find("img")
                        if img :
                            a_img = 1

                        a_tag2 = news2.find_all("a")[a_img]
                        title = a_tag2.get_text().strip()
                        link = a_tag2["href"]
                        press = news2.find("span", attrs={"class":"writing"}).get_text()

                        page_index_2 = page_index_2 + 1

                        ws1.append([page_index_2, title, link, press])

                page_index_2 = page_index_2 + index + 1

    page_number = 1
    paging_url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(page_number)
    paging_soup = create_paging_soup(paging_url)
    page_start = paging_soup.find("div", attrs={"class":"paging"}).find("strong").get_text()

    if(paging_soup.find("div", attrs={"class":"paging"}).find("a")) :
        page = paging_soup.find("div", attrs={"class":"paging"}).find_all("a")[-1]
        page_end = page.get_text()
    else :
        page_end = page_start

    if(page_end == "다음") :
        while(page_end == "다음") :
            page_number = page_number + 10
            paging_url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(page_number)
            paging_soup = create_paging_soup(paging_url)
            page = paging_soup.find("div", attrs={"class":"paging"}).find_all("a")[-1]
            page_end = page.get_text()

            if(page_end != "다음") :
                if(page_end == "이전") :
                    page_end = paging_soup.find("div", attrs={"class":"paging"}).find("strong").get_text()
                    scrape()
                else :
                    scrape()
    elif(page_end != "다음") :
        scrape()

def theme_102_249() :   # 사회 > 사건사고
    sid1_code = str(102)
    sid2_code = str(249)

    ws1 = wb.create_sheet("사회_사건사고")
    ws1 = wb["사회_사건사고"]
    ws1.append(["index", "제목", "링크", "언론사"])

    def scrape() :
        page_index_1 = 0
        page_index_2 = 10

        for i in range(int(page_start), int(page_end)+1) :
            url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(i)

            soup = create_soup(url)
            news_list1 = soup.find("ul", attrs={"class":"type06_headline"}).find_all("li")
            for index, news1 in enumerate(news_list1) :
                a_img = 0
                img = news1.find("img")
                if img :
                    a_img = 1

                a_tag1 = news1.find_all("a")[a_img]
                title = a_tag1.get_text().strip()
                link = a_tag1["href"]
                press = news1.find("span", attrs={"class":"writing"}).get_text()

                page_index_1 = page_index_1 + 1

                ws1.append([page_index_1, title, link, press])

            page_index_1 = page_index_1 + index + 1

            if(soup.find("ul", attrs={"class":"type06"})) :
                news_list2 = soup.find("ul", attrs={"class":"type06"}).find_all("li")
                for index, news2 in enumerate(news_list2) :
                        a_img = 0
                        img = news2.find("img")
                        if img :
                            a_img = 1

                        a_tag2 = news2.find_all("a")[a_img]
                        title = a_tag2.get_text().strip()
                        link = a_tag2["href"]
                        press = news2.find("span", attrs={"class":"writing"}).get_text()

                        page_index_2 = page_index_2 + 1

                        ws1.append([page_index_2, title, link, press])

                page_index_2 = page_index_2 + index + 1

    page_number = 1
    paging_url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(page_number)
    paging_soup = create_paging_soup(paging_url)
    page_start = paging_soup.find("div", attrs={"class":"paging"}).find("strong").get_text()

    if(paging_soup.find("div", attrs={"class":"paging"}).find("a")) :
        page = paging_soup.find("div", attrs={"class":"paging"}).find_all("a")[-1]
        page_end = page.get_text()
    else :
        page_end = page_start

    if(page_end == "다음") :
        while(page_end == "다음") :
            page_number = page_number + 10
            paging_url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(page_number)
            paging_soup = create_paging_soup(paging_url)
            page = paging_soup.find("div", attrs={"class":"paging"}).find_all("a")[-1]
            page_end = page.get_text()

            if(page_end != "다음") :
                if(page_end == "이전") :
                    page_end = paging_soup.find("div", attrs={"class":"paging"}).find("strong").get_text()
                    scrape()
                else :
                    scrape()
    elif(page_end != "다음") :
        scrape()

def theme_102_250() :   # 사회 > 교육
    sid1_code = str(102)
    sid2_code = str(250)

    ws1 = wb.create_sheet("사회_교육")
    ws1 = wb["사회_교육"]
    ws1.append(["index", "제목", "링크", "언론사"])

    def scrape() :
        page_index_1 = 0
        page_index_2 = 10

        for i in range(int(page_start), int(page_end)+1) :
            url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(i)

            soup = create_soup(url)
            news_list1 = soup.find("ul", attrs={"class":"type06_headline"}).find_all("li")
            for index, news1 in enumerate(news_list1) :
                a_img = 0
                img = news1.find("img")
                if img :
                    a_img = 1

                a_tag1 = news1.find_all("a")[a_img]
                title = a_tag1.get_text().strip()
                link = a_tag1["href"]
                press = news1.find("span", attrs={"class":"writing"}).get_text()

                page_index_1 = page_index_1 + 1

                ws1.append([page_index_1, title, link, press])

            page_index_1 = page_index_1 + index + 1

            if(soup.find("ul", attrs={"class":"type06"})) :
                news_list2 = soup.find("ul", attrs={"class":"type06"}).find_all("li")
                for index, news2 in enumerate(news_list2) :
                        a_img = 0
                        img = news2.find("img")
                        if img :
                            a_img = 1

                        a_tag2 = news2.find_all("a")[a_img]
                        title = a_tag2.get_text().strip()
                        link = a_tag2["href"]
                        press = news2.find("span", attrs={"class":"writing"}).get_text()

                        page_index_2 = page_index_2 + 1

                        ws1.append([page_index_2, title, link, press])

                page_index_2 = page_index_2 + index + 1

    page_number = 1
    paging_url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(page_number)
    paging_soup = create_paging_soup(paging_url)
    page_start = paging_soup.find("div", attrs={"class":"paging"}).find("strong").get_text()

    if(paging_soup.find("div", attrs={"class":"paging"}).find("a")) :
        page = paging_soup.find("div", attrs={"class":"paging"}).find_all("a")[-1]
        page_end = page.get_text()
    else :
        page_end = page_start

    if(page_end == "다음") :
        while(page_end == "다음") :
            page_number = page_number + 10
            paging_url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(page_number)
            paging_soup = create_paging_soup(paging_url)
            page = paging_soup.find("div", attrs={"class":"paging"}).find_all("a")[-1]
            page_end = page.get_text()

            if(page_end != "다음") :
                if(page_end == "이전") :
                    page_end = paging_soup.find("div", attrs={"class":"paging"}).find("strong").get_text()
                    scrape()
                else :
                    scrape()
    elif(page_end != "다음") :
        scrape()

def theme_102_251() :   # 사회 > 노동
    sid1_code = str(102)
    sid2_code = str(251)

    ws1 = wb.create_sheet("사회_노동")
    ws1 = wb["사회_노동"]
    ws1.append(["index", "제목", "링크", "언론사"])

    def scrape() :
        page_index_1 = 0
        page_index_2 = 10

        for i in range(int(page_start), int(page_end)+1) :
            url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(i)

            soup = create_soup(url)
            news_list1 = soup.find("ul", attrs={"class":"type06_headline"}).find_all("li")
            for index, news1 in enumerate(news_list1) :
                a_img = 0
                img = news1.find("img")
                if img :
                    a_img = 1

                a_tag1 = news1.find_all("a")[a_img]
                title = a_tag1.get_text().strip()
                link = a_tag1["href"]
                press = news1.find("span", attrs={"class":"writing"}).get_text()

                page_index_1 = page_index_1 + 1

                ws1.append([page_index_1, title, link, press])

            page_index_1 = page_index_1 + index + 1

            if(soup.find("ul", attrs={"class":"type06"})) :
                news_list2 = soup.find("ul", attrs={"class":"type06"}).find_all("li")
                for index, news2 in enumerate(news_list2) :
                        a_img = 0
                        img = news2.find("img")
                        if img :
                            a_img = 1

                        a_tag2 = news2.find_all("a")[a_img]
                        title = a_tag2.get_text().strip()
                        link = a_tag2["href"]
                        press = news2.find("span", attrs={"class":"writing"}).get_text()

                        page_index_2 = page_index_2 + 1

                        ws1.append([page_index_2, title, link, press])

                page_index_2 = page_index_2 + index + 1

    page_number = 1
    paging_url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(page_number)
    paging_soup = create_paging_soup(paging_url)
    page_start = paging_soup.find("div", attrs={"class":"paging"}).find("strong").get_text()

    if(paging_soup.find("div", attrs={"class":"paging"}).find("a")) :
        page = paging_soup.find("div", attrs={"class":"paging"}).find_all("a")[-1]
        page_end = page.get_text()
    else :
        page_end = page_start

    if(page_end == "다음") :
        while(page_end == "다음") :
            page_number = page_number + 10
            paging_url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(page_number)
            paging_soup = create_paging_soup(paging_url)
            page = paging_soup.find("div", attrs={"class":"paging"}).find_all("a")[-1]
            page_end = page.get_text()

            if(page_end != "다음") :
                if(page_end == "이전") :
                    page_end = paging_soup.find("div", attrs={"class":"paging"}).find("strong").get_text()
                    scrape()
                else :
                    scrape()
    elif(page_end != "다음") :
        scrape()

def theme_102_254() :   # 사회 > 언론
    sid1_code = str(102)
    sid2_code = str(254)

    ws1 = wb.create_sheet("사회_언론")
    ws1 = wb["사회_언론"]
    ws1.append(["index", "제목", "링크", "언론사"])

    def scrape() :
        page_index_1 = 0
        page_index_2 = 10

        for i in range(int(page_start), int(page_end)+1) :
            url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(i)

            soup = create_soup(url)
            news_list1 = soup.find("ul", attrs={"class":"type06_headline"}).find_all("li")
            for index, news1 in enumerate(news_list1) :
                a_img = 0
                img = news1.find("img")
                if img :
                    a_img = 1

                a_tag1 = news1.find_all("a")[a_img]
                title = a_tag1.get_text().strip()
                link = a_tag1["href"]
                press = news1.find("span", attrs={"class":"writing"}).get_text()

                page_index_1 = page_index_1 + 1

                ws1.append([page_index_1, title, link, press])

            page_index_1 = page_index_1 + index + 1

            if(soup.find("ul", attrs={"class":"type06"})) :
                news_list2 = soup.find("ul", attrs={"class":"type06"}).find_all("li")
                for index, news2 in enumerate(news_list2) :
                        a_img = 0
                        img = news2.find("img")
                        if img :
                            a_img = 1

                        a_tag2 = news2.find_all("a")[a_img]
                        title = a_tag2.get_text().strip()
                        link = a_tag2["href"]
                        press = news2.find("span", attrs={"class":"writing"}).get_text()

                        page_index_2 = page_index_2 + 1

                        ws1.append([page_index_2, title, link, press])

                page_index_2 = page_index_2 + index + 1

    page_number = 1
    paging_url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(page_number)
    paging_soup = create_paging_soup(paging_url)
    page_start = paging_soup.find("div", attrs={"class":"paging"}).find("strong").get_text()

    if(paging_soup.find("div", attrs={"class":"paging"}).find("a")) :
        page = paging_soup.find("div", attrs={"class":"paging"}).find_all("a")[-1]
        page_end = page.get_text()
    else :
        page_end = page_start

    if(page_end == "다음") :
        while(page_end == "다음") :
            page_number = page_number + 10
            paging_url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(page_number)
            paging_soup = create_paging_soup(paging_url)
            page = paging_soup.find("div", attrs={"class":"paging"}).find_all("a")[-1]
            page_end = page.get_text()

            if(page_end != "다음") :
                if(page_end == "이전") :
                    page_end = paging_soup.find("div", attrs={"class":"paging"}).find("strong").get_text()
                    scrape()
                else :
                    scrape()
    elif(page_end != "다음") :
        scrape()

def theme_102_252() :   # 사회 > 환경
    sid1_code = str(102)
    sid2_code = str(252)

    ws1 = wb.create_sheet("사회_환경")
    ws1 = wb["사회_환경"]
    ws1.append(["index", "제목", "링크", "언론사"])

    def scrape() :
        page_index_1 = 0
        page_index_2 = 10

        for i in range(int(page_start), int(page_end)+1) :
            url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(i)

            soup = create_soup(url)
            news_list1 = soup.find("ul", attrs={"class":"type06_headline"}).find_all("li")
            for index, news1 in enumerate(news_list1) :
                a_img = 0
                img = news1.find("img")
                if img :
                    a_img = 1

                a_tag1 = news1.find_all("a")[a_img]
                title = a_tag1.get_text().strip()
                link = a_tag1["href"]
                press = news1.find("span", attrs={"class":"writing"}).get_text()

                page_index_1 = page_index_1 + 1

                ws1.append([page_index_1, title, link, press])

            page_index_1 = page_index_1 + index + 1

            if(soup.find("ul", attrs={"class":"type06"})) :
                news_list2 = soup.find("ul", attrs={"class":"type06"}).find_all("li")
                for index, news2 in enumerate(news_list2) :
                        a_img = 0
                        img = news2.find("img")
                        if img :
                            a_img = 1

                        a_tag2 = news2.find_all("a")[a_img]
                        title = a_tag2.get_text().strip()
                        link = a_tag2["href"]
                        press = news2.find("span", attrs={"class":"writing"}).get_text()

                        page_index_2 = page_index_2 + 1

                        ws1.append([page_index_2, title, link, press])

                page_index_2 = page_index_2 + index + 1

    page_number = 1
    paging_url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(page_number)
    paging_soup = create_paging_soup(paging_url)
    page_start = paging_soup.find("div", attrs={"class":"paging"}).find("strong").get_text()

    if(paging_soup.find("div", attrs={"class":"paging"}).find("a")) :
        page = paging_soup.find("div", attrs={"class":"paging"}).find_all("a")[-1]
        page_end = page.get_text()
    else :
        page_end = page_start

    if(page_end == "다음") :
        while(page_end == "다음") :
            page_number = page_number + 10
            paging_url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(page_number)
            paging_soup = create_paging_soup(paging_url)
            page = paging_soup.find("div", attrs={"class":"paging"}).find_all("a")[-1]
            page_end = page.get_text()

            if(page_end != "다음") :
                if(page_end == "이전") :
                    page_end = paging_soup.find("div", attrs={"class":"paging"}).find("strong").get_text()
                    scrape()
                else :
                    scrape()
    elif(page_end != "다음") :
        scrape()

def theme_102_59b() :   # 사회 > 인권/복지
    sid1_code = str(102)
    sid2_code = "59b"   # 문자열

    ws1 = wb.create_sheet("사회_인권,복지")
    ws1 = wb["사회_인권,복지"]
    ws1.append(["index", "제목", "링크", "언론사"])

    def scrape() :
        page_index_1 = 0
        page_index_2 = 10

        for i in range(int(page_start), int(page_end)+1) :
            url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(i)

            soup = create_soup(url)
            news_list1 = soup.find("ul", attrs={"class":"type06_headline"}).find_all("li")
            for index, news1 in enumerate(news_list1) :
                a_img = 0
                img = news1.find("img")
                if img :
                    a_img = 1

                a_tag1 = news1.find_all("a")[a_img]
                title = a_tag1.get_text().strip()
                link = a_tag1["href"]
                press = news1.find("span", attrs={"class":"writing"}).get_text()

                page_index_1 = page_index_1 + 1

                ws1.append([page_index_1, title, link, press])

            page_index_1 = page_index_1 + index + 1

            if(soup.find("ul", attrs={"class":"type06"})) :
                news_list2 = soup.find("ul", attrs={"class":"type06"}).find_all("li")
                for index, news2 in enumerate(news_list2) :
                        a_img = 0
                        img = news2.find("img")
                        if img :
                            a_img = 1

                        a_tag2 = news2.find_all("a")[a_img]
                        title = a_tag2.get_text().strip()
                        link = a_tag2["href"]
                        press = news2.find("span", attrs={"class":"writing"}).get_text()

                        page_index_2 = page_index_2 + 1

                        ws1.append([page_index_2, title, link, press])

                page_index_2 = page_index_2 + index + 1

    page_number = 1
    paging_url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(page_number)
    paging_soup = create_paging_soup(paging_url)
    page_start = paging_soup.find("div", attrs={"class":"paging"}).find("strong").get_text()

    if(paging_soup.find("div", attrs={"class":"paging"}).find("a")) :
        page = paging_soup.find("div", attrs={"class":"paging"}).find_all("a")[-1]
        page_end = page.get_text()
    else :
        page_end = page_start

    if(page_end == "다음") :
        while(page_end == "다음") :
            page_number = page_number + 10
            paging_url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(page_number)
            paging_soup = create_paging_soup(paging_url)
            page = paging_soup.find("div", attrs={"class":"paging"}).find_all("a")[-1]
            page_end = page.get_text()

            if(page_end != "다음") :
                if(page_end == "이전") :
                    page_end = paging_soup.find("div", attrs={"class":"paging"}).find("strong").get_text()
                    scrape()
                else :
                    scrape()
    elif(page_end != "다음") :
        scrape()

def theme_102_255() :   # 사회 > 식품/의료
    sid1_code = str(102)
    sid2_code = str(255)

    ws1 = wb.create_sheet("사회_식품,의료")
    ws1 = wb["사회_식품,의료"]
    ws1.append(["index", "제목", "링크", "언론사"])

    def scrape() :
        page_index_1 = 0
        page_index_2 = 10

        for i in range(int(page_start), int(page_end)+1) :
            url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(i)

            soup = create_soup(url)
            news_list1 = soup.find("ul", attrs={"class":"type06_headline"}).find_all("li")
            for index, news1 in enumerate(news_list1) :
                a_img = 0
                img = news1.find("img")
                if img :
                    a_img = 1

                a_tag1 = news1.find_all("a")[a_img]
                title = a_tag1.get_text().strip()
                link = a_tag1["href"]
                press = news1.find("span", attrs={"class":"writing"}).get_text()

                page_index_1 = page_index_1 + 1

                ws1.append([page_index_1, title, link, press])

            page_index_1 = page_index_1 + index + 1

            if(soup.find("ul", attrs={"class":"type06"})) :
                news_list2 = soup.find("ul", attrs={"class":"type06"}).find_all("li")
                for index, news2 in enumerate(news_list2) :
                        a_img = 0
                        img = news2.find("img")
                        if img :
                            a_img = 1

                        a_tag2 = news2.find_all("a")[a_img]
                        title = a_tag2.get_text().strip()
                        link = a_tag2["href"]
                        press = news2.find("span", attrs={"class":"writing"}).get_text()

                        page_index_2 = page_index_2 + 1

                        ws1.append([page_index_2, title, link, press])

                page_index_2 = page_index_2 + index + 1

    page_number = 1
    paging_url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(page_number)
    paging_soup = create_paging_soup(paging_url)
    page_start = paging_soup.find("div", attrs={"class":"paging"}).find("strong").get_text()

    if(paging_soup.find("div", attrs={"class":"paging"}).find("a")) :
        page = paging_soup.find("div", attrs={"class":"paging"}).find_all("a")[-1]
        page_end = page.get_text()
    else :
        page_end = page_start

    if(page_end == "다음") :
        while(page_end == "다음") :
            page_number = page_number + 10
            paging_url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(page_number)
            paging_soup = create_paging_soup(paging_url)
            page = paging_soup.find("div", attrs={"class":"paging"}).find_all("a")[-1]
            page_end = page.get_text()

            if(page_end != "다음") :
                if(page_end == "이전") :
                    page_end = paging_soup.find("div", attrs={"class":"paging"}).find("strong").get_text()
                    scrape()
                else :
                    scrape()
    elif(page_end != "다음") :
        scrape()

def theme_102_256() :   # 사회 > 지역
    sid1_code = str(102)
    sid2_code = str(256)

    ws1 = wb.create_sheet("사회_지역")
    ws1 = wb["사회_지역"]
    ws1.append(["index", "제목", "링크", "언론사"])

    def scrape() :
        page_index_1 = 0
        page_index_2 = 10

        for i in range(int(page_start), int(page_end)+1) :
            url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(i)

            soup = create_soup(url)
            news_list1 = soup.find("ul", attrs={"class":"type06_headline"}).find_all("li")
            for index, news1 in enumerate(news_list1) :
                a_img = 0
                img = news1.find("img")
                if img :
                    a_img = 1

                a_tag1 = news1.find_all("a")[a_img]
                title = a_tag1.get_text().strip()
                link = a_tag1["href"]
                press = news1.find("span", attrs={"class":"writing"}).get_text()

                page_index_1 = page_index_1 + 1

                ws1.append([page_index_1, title, link, press])

            page_index_1 = page_index_1 + index + 1

            if(soup.find("ul", attrs={"class":"type06"})) :
                news_list2 = soup.find("ul", attrs={"class":"type06"}).find_all("li")
                for index, news2 in enumerate(news_list2) :
                        a_img = 0
                        img = news2.find("img")
                        if img :
                            a_img = 1

                        a_tag2 = news2.find_all("a")[a_img]
                        title = a_tag2.get_text().strip()
                        link = a_tag2["href"]
                        press = news2.find("span", attrs={"class":"writing"}).get_text()

                        page_index_2 = page_index_2 + 1

                        ws1.append([page_index_2, title, link, press])

                page_index_2 = page_index_2 + index + 1

    page_number = 1
    paging_url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(page_number)
    paging_soup = create_paging_soup(paging_url)
    page_start = paging_soup.find("div", attrs={"class":"paging"}).find("strong").get_text()

    if(paging_soup.find("div", attrs={"class":"paging"}).find("a")) :
        page = paging_soup.find("div", attrs={"class":"paging"}).find_all("a")[-1]
        page_end = page.get_text()
    else :
        page_end = page_start

    if(page_end == "다음") :
        while(page_end == "다음") :
            page_number = page_number + 10
            paging_url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(page_number)
            paging_soup = create_paging_soup(paging_url)
            page = paging_soup.find("div", attrs={"class":"paging"}).find_all("a")[-1]
            page_end = page.get_text()

            if(page_end != "다음") :
                if(page_end == "이전") :
                    page_end = paging_soup.find("div", attrs={"class":"paging"}).find("strong").get_text()
                    scrape()
                else :
                    scrape()
    elif(page_end != "다음") :
        scrape()

def theme_102_276() :   # 사회 > 인물
    sid1_code = str(102)
    sid2_code = str(276)

    ws1 = wb.create_sheet("사회_인물")
    ws1 = wb["사회_인물"]
    ws1.append(["index", "제목", "링크", "언론사"])

    def scrape() :
        page_index_1 = 0
        page_index_2 = 10

        for i in range(int(page_start), int(page_end)+1) :
            url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(i)

            soup = create_soup(url)
            news_list1 = soup.find("ul", attrs={"class":"type06_headline"}).find_all("li")
            for index, news1 in enumerate(news_list1) :
                a_img = 0
                img = news1.find("img")
                if img :
                    a_img = 1

                a_tag1 = news1.find_all("a")[a_img]
                title = a_tag1.get_text().strip()
                link = a_tag1["href"]
                press = news1.find("span", attrs={"class":"writing"}).get_text()

                page_index_1 = page_index_1 + 1

                ws1.append([page_index_1, title, link, press])

            page_index_1 = page_index_1 + index + 1

            if(soup.find("ul", attrs={"class":"type06"})) :
                news_list2 = soup.find("ul", attrs={"class":"type06"}).find_all("li")
                for index, news2 in enumerate(news_list2) :
                        a_img = 0
                        img = news2.find("img")
                        if img :
                            a_img = 1

                        a_tag2 = news2.find_all("a")[a_img]
                        title = a_tag2.get_text().strip()
                        link = a_tag2["href"]
                        press = news2.find("span", attrs={"class":"writing"}).get_text()

                        page_index_2 = page_index_2 + 1

                        ws1.append([page_index_2, title, link, press])

                page_index_2 = page_index_2 + index + 1

    page_number = 1
    paging_url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(page_number)
    paging_soup = create_paging_soup(paging_url)
    page_start = paging_soup.find("div", attrs={"class":"paging"}).find("strong").get_text()

    if(paging_soup.find("div", attrs={"class":"paging"}).find("a")) :
        page = paging_soup.find("div", attrs={"class":"paging"}).find_all("a")[-1]
        page_end = page.get_text()
    else :
        page_end = page_start

    if(page_end == "다음") :
        while(page_end == "다음") :
            page_number = page_number + 10
            paging_url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(page_number)
            paging_soup = create_paging_soup(paging_url)
            page = paging_soup.find("div", attrs={"class":"paging"}).find_all("a")[-1]
            page_end = page.get_text()

            if(page_end != "다음") :
                if(page_end == "이전") :
                    page_end = paging_soup.find("div", attrs={"class":"paging"}).find("strong").get_text()
                    scrape()
                else :
                    scrape()
    elif(page_end != "다음") :
        scrape()

def theme_102_257() :   # 사회 > 사회일반
    sid1_code = str(102)
    sid2_code = str(257)

    ws1 = wb.create_sheet("사회_사회일반")
    ws1 = wb["사회_사회일반"]
    ws1.append(["index", "제목", "링크", "언론사"])

    def scrape() :
        page_index_1 = 0
        page_index_2 = 10

        for i in range(int(page_start), int(page_end)+1) :
            url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(i)

            soup = create_soup(url)
            news_list1 = soup.find("ul", attrs={"class":"type06_headline"}).find_all("li")
            for index, news1 in enumerate(news_list1) :
                a_img = 0
                img = news1.find("img")
                if img :
                    a_img = 1

                a_tag1 = news1.find_all("a")[a_img]
                title = a_tag1.get_text().strip()
                link = a_tag1["href"]
                press = news1.find("span", attrs={"class":"writing"}).get_text()

                page_index_1 = page_index_1 + 1

                ws1.append([page_index_1, title, link, press])

            page_index_1 = page_index_1 + index + 1

            if(soup.find("ul", attrs={"class":"type06"})) :
                news_list2 = soup.find("ul", attrs={"class":"type06"}).find_all("li")
                for index, news2 in enumerate(news_list2) :
                        a_img = 0
                        img = news2.find("img")
                        if img :
                            a_img = 1

                        a_tag2 = news2.find_all("a")[a_img]
                        title = a_tag2.get_text().strip()
                        link = a_tag2["href"]
                        press = news2.find("span", attrs={"class":"writing"}).get_text()

                        page_index_2 = page_index_2 + 1

                        ws1.append([page_index_2, title, link, press])

                page_index_2 = page_index_2 + index + 1

    page_number = 1
    paging_url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(page_number)
    paging_soup = create_paging_soup(paging_url)
    page_start = paging_soup.find("div", attrs={"class":"paging"}).find("strong").get_text()

    if(paging_soup.find("div", attrs={"class":"paging"}).find("a")) :
        page = paging_soup.find("div", attrs={"class":"paging"}).find_all("a")[-1]
        page_end = page.get_text()
    else :
        page_end = page_start

    if(page_end == "다음") :
        while(page_end == "다음") :
            page_number = page_number + 10
            paging_url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(page_number)
            paging_soup = create_paging_soup(paging_url)
            page = paging_soup.find("div", attrs={"class":"paging"}).find_all("a")[-1]
            page_end = page.get_text()

            if(page_end != "다음") :
                if(page_end == "이전") :
                    page_end = paging_soup.find("div", attrs={"class":"paging"}).find("strong").get_text()
                    scrape()
                else :
                    scrape()
    elif(page_end != "다음") :
        scrape()

def theme_103_241() :   # 생활/문화 > 건강정보
    sid1_code = str(103)
    sid2_code = str(241)

    ws1 = wb.create_sheet("생활,문화_건강정보")
    ws1 = wb["생활,문화_건강정보"]
    ws1.append(["index", "제목", "링크", "언론사"])

    def scrape() :
        page_index_1 = 0
        page_index_2 = 10

        for i in range(int(page_start), int(page_end)+1) :
            url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(i)

            soup = create_soup(url)
            news_list1 = soup.find("ul", attrs={"class":"type06_headline"}).find_all("li")
            for index, news1 in enumerate(news_list1) :
                a_img = 0
                img = news1.find("img")
                if img :
                    a_img = 1

                a_tag1 = news1.find_all("a")[a_img]
                title = a_tag1.get_text().strip()
                link = a_tag1["href"]
                press = news1.find("span", attrs={"class":"writing"}).get_text()

                page_index_1 = page_index_1 + 1

                ws1.append([page_index_1, title, link, press])

            page_index_1 = page_index_1 + index + 1

            if(soup.find("ul", attrs={"class":"type06"})) :
                news_list2 = soup.find("ul", attrs={"class":"type06"}).find_all("li")
                for index, news2 in enumerate(news_list2) :
                        a_img = 0
                        img = news2.find("img")
                        if img :
                            a_img = 1

                        a_tag2 = news2.find_all("a")[a_img]
                        title = a_tag2.get_text().strip()
                        link = a_tag2["href"]
                        press = news2.find("span", attrs={"class":"writing"}).get_text()

                        page_index_2 = page_index_2 + 1

                        ws1.append([page_index_2, title, link, press])

                page_index_2 = page_index_2 + index + 1

    page_number = 1
    paging_url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(page_number)
    paging_soup = create_paging_soup(paging_url)
    page_start = paging_soup.find("div", attrs={"class":"paging"}).find("strong").get_text()

    if(paging_soup.find("div", attrs={"class":"paging"}).find("a")) :
        page = paging_soup.find("div", attrs={"class":"paging"}).find_all("a")[-1]
        page_end = page.get_text()
    else :
        page_end = page_start

    if(page_end == "다음") :
        while(page_end == "다음") :
            page_number = page_number + 10
            paging_url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(page_number)
            paging_soup = create_paging_soup(paging_url)
            page = paging_soup.find("div", attrs={"class":"paging"}).find_all("a")[-1]
            page_end = page.get_text()

            if(page_end != "다음") :
                if(page_end == "이전") :
                    page_end = paging_soup.find("div", attrs={"class":"paging"}).find("strong").get_text()
                    scrape()
                else :
                    scrape()
    elif(page_end != "다음") :
        scrape()

def theme_103_239() :   # 생활/문화 > 자동차/시승기
    sid1_code = str(103)
    sid2_code = str(239)

    ws1 = wb.create_sheet("생활,문화_자동차,시승기")
    ws1 = wb["생활,문화_자동차,시승기"]
    ws1.append(["index", "제목", "링크", "언론사"])

    def scrape() :
        page_index_1 = 0
        page_index_2 = 10

        for i in range(int(page_start), int(page_end)+1) :
            url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(i)

            soup = create_soup(url)
            news_list1 = soup.find("ul", attrs={"class":"type06_headline"}).find_all("li")
            for index, news1 in enumerate(news_list1) :
                a_img = 0
                img = news1.find("img")
                if img :
                    a_img = 1

                a_tag1 = news1.find_all("a")[a_img]
                title = a_tag1.get_text().strip()
                link = a_tag1["href"]
                press = news1.find("span", attrs={"class":"writing"}).get_text()

                page_index_1 = page_index_1 + 1

                ws1.append([page_index_1, title, link, press])

            page_index_1 = page_index_1 + index + 1

            if(soup.find("ul", attrs={"class":"type06"})) :
                news_list2 = soup.find("ul", attrs={"class":"type06"}).find_all("li")
                for index, news2 in enumerate(news_list2) :
                        a_img = 0
                        img = news2.find("img")
                        if img :
                            a_img = 1

                        a_tag2 = news2.find_all("a")[a_img]
                        title = a_tag2.get_text().strip()
                        link = a_tag2["href"]
                        press = news2.find("span", attrs={"class":"writing"}).get_text()

                        page_index_2 = page_index_2 + 1

                        ws1.append([page_index_2, title, link, press])

                page_index_2 = page_index_2 + index + 1

    page_number = 1
    paging_url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(page_number)
    paging_soup = create_paging_soup(paging_url)
    page_start = paging_soup.find("div", attrs={"class":"paging"}).find("strong").get_text()

    if(paging_soup.find("div", attrs={"class":"paging"}).find("a")) :
        page = paging_soup.find("div", attrs={"class":"paging"}).find_all("a")[-1]
        page_end = page.get_text()
    else :
        page_end = page_start

    if(page_end == "다음") :
        while(page_end == "다음") :
            page_number = page_number + 10
            paging_url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(page_number)
            paging_soup = create_paging_soup(paging_url)
            page = paging_soup.find("div", attrs={"class":"paging"}).find_all("a")[-1]
            page_end = page.get_text()

            if(page_end != "다음") :
                if(page_end == "이전") :
                    page_end = paging_soup.find("div", attrs={"class":"paging"}).find("strong").get_text()
                    scrape()
                else :
                    scrape()
    elif(page_end != "다음") :
        scrape()

def theme_103_240() :   # 생활/문화 > 도로/교통
    sid1_code = str(103)
    sid2_code = str(240)

    ws1 = wb.create_sheet("생활,문화_도로,교통")
    ws1 = wb["생활,문화_도로,교통"]
    ws1.append(["index", "제목", "링크", "언론사"])

    def scrape() :
        page_index_1 = 0
        page_index_2 = 10

        for i in range(int(page_start), int(page_end)+1) :
            url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(i)

            soup = create_soup(url)
            news_list1 = soup.find("ul", attrs={"class":"type06_headline"}).find_all("li")
            for index, news1 in enumerate(news_list1) :
                a_img = 0
                img = news1.find("img")
                if img :
                    a_img = 1

                a_tag1 = news1.find_all("a")[a_img]
                title = a_tag1.get_text().strip()
                link = a_tag1["href"]
                press = news1.find("span", attrs={"class":"writing"}).get_text()

                page_index_1 = page_index_1 + 1

                ws1.append([page_index_1, title, link, press])

            page_index_1 = page_index_1 + index + 1

            if(soup.find("ul", attrs={"class":"type06"})) :
                news_list2 = soup.find("ul", attrs={"class":"type06"}).find_all("li")
                for index, news2 in enumerate(news_list2) :
                        a_img = 0
                        img = news2.find("img")
                        if img :
                            a_img = 1

                        a_tag2 = news2.find_all("a")[a_img]
                        title = a_tag2.get_text().strip()
                        link = a_tag2["href"]
                        press = news2.find("span", attrs={"class":"writing"}).get_text()

                        page_index_2 = page_index_2 + 1

                        ws1.append([page_index_2, title, link, press])

                page_index_2 = page_index_2 + index + 1

    page_number = 1
    paging_url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(page_number)
    paging_soup = create_paging_soup(paging_url)
    page_start = paging_soup.find("div", attrs={"class":"paging"}).find("strong").get_text()

    if(paging_soup.find("div", attrs={"class":"paging"}).find("a")) :
        page = paging_soup.find("div", attrs={"class":"paging"}).find_all("a")[-1]
        page_end = page.get_text()
    else :
        page_end = page_start

    if(page_end == "다음") :
        while(page_end == "다음") :
            page_number = page_number + 10
            paging_url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(page_number)
            paging_soup = create_paging_soup(paging_url)
            page = paging_soup.find("div", attrs={"class":"paging"}).find_all("a")[-1]
            page_end = page.get_text()

            if(page_end != "다음") :
                if(page_end == "이전") :
                    page_end = paging_soup.find("div", attrs={"class":"paging"}).find("strong").get_text()
                    scrape()
                else :
                    scrape()
    elif(page_end != "다음") :
        scrape()

def theme_103_237() :   # 생활/문화 > 여행/레저
    sid1_code = str(103)
    sid2_code = str(237)

    ws1 = wb.create_sheet("생활,문화_여행,레저")
    ws1 = wb["생활,문화_여행,레저"]
    ws1.append(["index", "제목", "링크", "언론사"])

    def scrape() :
        page_index_1 = 0
        page_index_2 = 10

        for i in range(int(page_start), int(page_end)+1) :
            url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(i)

            soup = create_soup(url)
            news_list1 = soup.find("ul", attrs={"class":"type06_headline"}).find_all("li")
            for index, news1 in enumerate(news_list1) :
                a_img = 0
                img = news1.find("img")
                if img :
                    a_img = 1

                a_tag1 = news1.find_all("a")[a_img]
                title = a_tag1.get_text().strip()
                link = a_tag1["href"]
                press = news1.find("span", attrs={"class":"writing"}).get_text()

                page_index_1 = page_index_1 + 1

                ws1.append([page_index_1, title, link, press])

            page_index_1 = page_index_1 + index + 1

            if(soup.find("ul", attrs={"class":"type06"})) :
                news_list2 = soup.find("ul", attrs={"class":"type06"}).find_all("li")
                for index, news2 in enumerate(news_list2) :
                        a_img = 0
                        img = news2.find("img")
                        if img :
                            a_img = 1

                        a_tag2 = news2.find_all("a")[a_img]
                        title = a_tag2.get_text().strip()
                        link = a_tag2["href"]
                        press = news2.find("span", attrs={"class":"writing"}).get_text()

                        page_index_2 = page_index_2 + 1

                        ws1.append([page_index_2, title, link, press])

                page_index_2 = page_index_2 + index + 1

    page_number = 1
    paging_url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(page_number)
    paging_soup = create_paging_soup(paging_url)
    page_start = paging_soup.find("div", attrs={"class":"paging"}).find("strong").get_text()

    if(paging_soup.find("div", attrs={"class":"paging"}).find("a")) :
        page = paging_soup.find("div", attrs={"class":"paging"}).find_all("a")[-1]
        page_end = page.get_text()
    else :
        page_end = page_start

    if(page_end == "다음") :
        while(page_end == "다음") :
            page_number = page_number + 10
            paging_url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(page_number)
            paging_soup = create_paging_soup(paging_url)
            page = paging_soup.find("div", attrs={"class":"paging"}).find_all("a")[-1]
            page_end = page.get_text()

            if(page_end != "다음") :
                if(page_end == "이전") :
                    page_end = paging_soup.find("div", attrs={"class":"paging"}).find("strong").get_text()
                    scrape()
                else :
                    scrape()
    elif(page_end != "다음") :
        scrape()

def theme_103_238() :   # 생활/문화 > 음식/맛집
    sid1_code = str(103)
    sid2_code = str(238)

    ws1 = wb.create_sheet("생활,문화_음식,맛집")
    ws1 = wb["생활,문화_음식,맛집"]
    ws1.append(["index", "제목", "링크", "언론사"])

    def scrape() :
        page_index_1 = 0
        page_index_2 = 10

        for i in range(int(page_start), int(page_end)+1) :
            url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(i)

            soup = create_soup(url)
            news_list1 = soup.find("ul", attrs={"class":"type06_headline"}).find_all("li")
            for index, news1 in enumerate(news_list1) :
                a_img = 0
                img = news1.find("img")
                if img :
                    a_img = 1

                a_tag1 = news1.find_all("a")[a_img]
                title = a_tag1.get_text().strip()
                link = a_tag1["href"]
                press = news1.find("span", attrs={"class":"writing"}).get_text()

                page_index_1 = page_index_1 + 1

                ws1.append([page_index_1, title, link, press])

            page_index_1 = page_index_1 + index + 1

            if(soup.find("ul", attrs={"class":"type06"})) :
                news_list2 = soup.find("ul", attrs={"class":"type06"}).find_all("li")
                for index, news2 in enumerate(news_list2) :
                        a_img = 0
                        img = news2.find("img")
                        if img :
                            a_img = 1

                        a_tag2 = news2.find_all("a")[a_img]
                        title = a_tag2.get_text().strip()
                        link = a_tag2["href"]
                        press = news2.find("span", attrs={"class":"writing"}).get_text()

                        page_index_2 = page_index_2 + 1

                        ws1.append([page_index_2, title, link, press])

                page_index_2 = page_index_2 + index + 1

    page_number = 1
    paging_url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(page_number)
    paging_soup = create_paging_soup(paging_url)
    page_start = paging_soup.find("div", attrs={"class":"paging"}).find("strong").get_text()

    if(paging_soup.find("div", attrs={"class":"paging"}).find("a")) :
        page = paging_soup.find("div", attrs={"class":"paging"}).find_all("a")[-1]
        page_end = page.get_text()
    else :
        page_end = page_start

    if(page_end == "다음") :
        while(page_end == "다음") :
            page_number = page_number + 10
            paging_url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(page_number)
            paging_soup = create_paging_soup(paging_url)
            page = paging_soup.find("div", attrs={"class":"paging"}).find_all("a")[-1]
            page_end = page.get_text()

            if(page_end != "다음") :
                if(page_end == "이전") :
                    page_end = paging_soup.find("div", attrs={"class":"paging"}).find("strong").get_text()
                    scrape()
                else :
                    scrape()
    elif(page_end != "다음") :
        scrape()

def theme_103_376() :   # 생활/문화 > 패션/뷰티
    sid1_code = str(103)
    sid2_code = str(376)

    ws1 = wb.create_sheet("생활,문화_패션,뷰티")
    ws1 = wb["생활,문화_패션,뷰티"]
    ws1.append(["index", "제목", "링크", "언론사"])

    def scrape() :
        page_index_1 = 0
        page_index_2 = 10

        for i in range(int(page_start), int(page_end)+1) :
            url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(i)

            soup = create_soup(url)
            news_list1 = soup.find("ul", attrs={"class":"type06_headline"}).find_all("li")
            for index, news1 in enumerate(news_list1) :
                a_img = 0
                img = news1.find("img")
                if img :
                    a_img = 1

                a_tag1 = news1.find_all("a")[a_img]
                title = a_tag1.get_text().strip()
                link = a_tag1["href"]
                press = news1.find("span", attrs={"class":"writing"}).get_text()

                page_index_1 = page_index_1 + 1

                ws1.append([page_index_1, title, link, press])

            page_index_1 = page_index_1 + index + 1

            if(soup.find("ul", attrs={"class":"type06"})) :
                news_list2 = soup.find("ul", attrs={"class":"type06"}).find_all("li")
                for index, news2 in enumerate(news_list2) :
                        a_img = 0
                        img = news2.find("img")
                        if img :
                            a_img = 1

                        a_tag2 = news2.find_all("a")[a_img]
                        title = a_tag2.get_text().strip()
                        link = a_tag2["href"]
                        press = news2.find("span", attrs={"class":"writing"}).get_text()

                        page_index_2 = page_index_2 + 1

                        ws1.append([page_index_2, title, link, press])

                page_index_2 = page_index_2 + index + 1

    page_number = 1
    paging_url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(page_number)
    paging_soup = create_paging_soup(paging_url)
    page_start = paging_soup.find("div", attrs={"class":"paging"}).find("strong").get_text()

    if(paging_soup.find("div", attrs={"class":"paging"}).find("a")) :
        page = paging_soup.find("div", attrs={"class":"paging"}).find_all("a")[-1]
        page_end = page.get_text()
    else :
        page_end = page_start

    if(page_end == "다음") :
        while(page_end == "다음") :
            page_number = page_number + 10
            paging_url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(page_number)
            paging_soup = create_paging_soup(paging_url)
            page = paging_soup.find("div", attrs={"class":"paging"}).find_all("a")[-1]
            page_end = page.get_text()

            if(page_end != "다음") :
                if(page_end == "이전") :
                    page_end = paging_soup.find("div", attrs={"class":"paging"}).find("strong").get_text()
                    scrape()
                else :
                    scrape()
    elif(page_end != "다음") :
        scrape()

def theme_103_242() :   # 생활/문화 > 공연/전시
    sid1_code = str(103)
    sid2_code = str(242)

    ws1 = wb.create_sheet("생활,문화_공연,전시")
    ws1 = wb["생활,문화_공연,전시"]
    ws1.append(["index", "제목", "링크", "언론사"])

    def scrape() :
        page_index_1 = 0
        page_index_2 = 10

        for i in range(int(page_start), int(page_end)+1) :
            url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(i)

            soup = create_soup(url)
            news_list1 = soup.find("ul", attrs={"class":"type06_headline"}).find_all("li")
            for index, news1 in enumerate(news_list1) :
                a_img = 0
                img = news1.find("img")
                if img :
                    a_img = 1

                a_tag1 = news1.find_all("a")[a_img]
                title = a_tag1.get_text().strip()
                link = a_tag1["href"]
                press = news1.find("span", attrs={"class":"writing"}).get_text()

                page_index_1 = page_index_1 + 1

                ws1.append([page_index_1, title, link, press])

            page_index_1 = page_index_1 + index + 1

            if(soup.find("ul", attrs={"class":"type06"})) :
                news_list2 = soup.find("ul", attrs={"class":"type06"}).find_all("li")
                for index, news2 in enumerate(news_list2) :
                        a_img = 0
                        img = news2.find("img")
                        if img :
                            a_img = 1

                        a_tag2 = news2.find_all("a")[a_img]
                        title = a_tag2.get_text().strip()
                        link = a_tag2["href"]
                        press = news2.find("span", attrs={"class":"writing"}).get_text()

                        page_index_2 = page_index_2 + 1

                        ws1.append([page_index_2, title, link, press])

                page_index_2 = page_index_2 + index + 1

    page_number = 1
    paging_url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(page_number)
    paging_soup = create_paging_soup(paging_url)
    page_start = paging_soup.find("div", attrs={"class":"paging"}).find("strong").get_text()

    if(paging_soup.find("div", attrs={"class":"paging"}).find("a")) :
        page = paging_soup.find("div", attrs={"class":"paging"}).find_all("a")[-1]
        page_end = page.get_text()
    else :
        page_end = page_start

    if(page_end == "다음") :
        while(page_end == "다음") :
            page_number = page_number + 10
            paging_url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(page_number)
            paging_soup = create_paging_soup(paging_url)
            page = paging_soup.find("div", attrs={"class":"paging"}).find_all("a")[-1]
            page_end = page.get_text()

            if(page_end != "다음") :
                if(page_end == "이전") :
                    page_end = paging_soup.find("div", attrs={"class":"paging"}).find("strong").get_text()
                    scrape()
                else :
                    scrape()
    elif(page_end != "다음") :
        scrape()

def theme_103_243() :   # 생활/문화 > 책
    sid1_code = str(103)
    sid2_code = str(243)

    ws1 = wb.create_sheet("생활,문화_책")
    ws1 = wb["생활,문화_책"]
    ws1.append(["index", "제목", "링크", "언론사"])

    def scrape() :
        page_index_1 = 0
        page_index_2 = 10

        for i in range(int(page_start), int(page_end)+1) :
            url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(i)

            soup = create_soup(url)
            news_list1 = soup.find("ul", attrs={"class":"type06_headline"}).find_all("li")
            for index, news1 in enumerate(news_list1) :
                a_img = 0
                img = news1.find("img")
                if img :
                    a_img = 1

                a_tag1 = news1.find_all("a")[a_img]
                title = a_tag1.get_text().strip()
                link = a_tag1["href"]
                press = news1.find("span", attrs={"class":"writing"}).get_text()

                page_index_1 = page_index_1 + 1

                ws1.append([page_index_1, title, link, press])

            page_index_1 = page_index_1 + index + 1

            if(soup.find("ul", attrs={"class":"type06"})) :
                news_list2 = soup.find("ul", attrs={"class":"type06"}).find_all("li")
                for index, news2 in enumerate(news_list2) :
                        a_img = 0
                        img = news2.find("img")
                        if img :
                            a_img = 1

                        a_tag2 = news2.find_all("a")[a_img]
                        title = a_tag2.get_text().strip()
                        link = a_tag2["href"]
                        press = news2.find("span", attrs={"class":"writing"}).get_text()

                        page_index_2 = page_index_2 + 1

                        ws1.append([page_index_2, title, link, press])

                page_index_2 = page_index_2 + index + 1

    page_number = 1
    paging_url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(page_number)
    paging_soup = create_paging_soup(paging_url)
    page_start = paging_soup.find("div", attrs={"class":"paging"}).find("strong").get_text()

    if(paging_soup.find("div", attrs={"class":"paging"}).find("a")) :
        page = paging_soup.find("div", attrs={"class":"paging"}).find_all("a")[-1]
        page_end = page.get_text()
    else :
        page_end = page_start

    if(page_end == "다음") :
        while(page_end == "다음") :
            page_number = page_number + 10
            paging_url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(page_number)
            paging_soup = create_paging_soup(paging_url)
            page = paging_soup.find("div", attrs={"class":"paging"}).find_all("a")[-1]
            page_end = page.get_text()

            if(page_end != "다음") :
                if(page_end == "이전") :
                    page_end = paging_soup.find("div", attrs={"class":"paging"}).find("strong").get_text()
                    scrape()
                else :
                    scrape()
    elif(page_end != "다음") :
        scrape()

def theme_103_244() :   # 생활/문화 > 종교
    sid1_code = str(103)
    sid2_code = str(244)

    ws1 = wb.create_sheet("생활,문화_종교")
    ws1 = wb["생활,문화_종교"]
    ws1.append(["index", "제목", "링크", "언론사"])

    def scrape() :
        page_index_1 = 0
        page_index_2 = 10

        for i in range(int(page_start), int(page_end)+1) :
            url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(i)

            soup = create_soup(url)
            news_list1 = soup.find("ul", attrs={"class":"type06_headline"}).find_all("li")
            for index, news1 in enumerate(news_list1) :
                a_img = 0
                img = news1.find("img")
                if img :
                    a_img = 1

                a_tag1 = news1.find_all("a")[a_img]
                title = a_tag1.get_text().strip()
                link = a_tag1["href"]
                press = news1.find("span", attrs={"class":"writing"}).get_text()

                page_index_1 = page_index_1 + 1

                ws1.append([page_index_1, title, link, press])

            page_index_1 = page_index_1 + index + 1

            if(soup.find("ul", attrs={"class":"type06"})) :
                news_list2 = soup.find("ul", attrs={"class":"type06"}).find_all("li")
                for index, news2 in enumerate(news_list2) :
                        a_img = 0
                        img = news2.find("img")
                        if img :
                            a_img = 1

                        a_tag2 = news2.find_all("a")[a_img]
                        title = a_tag2.get_text().strip()
                        link = a_tag2["href"]
                        press = news2.find("span", attrs={"class":"writing"}).get_text()

                        page_index_2 = page_index_2 + 1

                        ws1.append([page_index_2, title, link, press])

                page_index_2 = page_index_2 + index + 1

    page_number = 1
    paging_url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(page_number)
    paging_soup = create_paging_soup(paging_url)
    page_start = paging_soup.find("div", attrs={"class":"paging"}).find("strong").get_text()

    if(paging_soup.find("div", attrs={"class":"paging"}).find("a")) :
        page = paging_soup.find("div", attrs={"class":"paging"}).find_all("a")[-1]
        page_end = page.get_text()
    else :
        page_end = page_start

    if(page_end == "다음") :
        while(page_end == "다음") :
            page_number = page_number + 10
            paging_url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(page_number)
            paging_soup = create_paging_soup(paging_url)
            page = paging_soup.find("div", attrs={"class":"paging"}).find_all("a")[-1]
            page_end = page.get_text()

            if(page_end != "다음") :
                if(page_end == "이전") :
                    page_end = paging_soup.find("div", attrs={"class":"paging"}).find("strong").get_text()
                    scrape()
                else :
                    scrape()
    elif(page_end != "다음") :
        scrape()

def theme_103_248() :   # 생활/문화 > 날씨
    sid1_code = str(103)
    sid2_code = str(248)

    ws1 = wb.create_sheet("생활,문화_날씨")
    ws1 = wb["생활,문화_날씨"]
    ws1.append(["index", "제목", "링크", "언론사"])

    def scrape() :
        page_index_1 = 0
        page_index_2 = 10

        for i in range(int(page_start), int(page_end)+1) :
            url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(i)

            soup = create_soup(url)
            news_list1 = soup.find("ul", attrs={"class":"type06_headline"}).find_all("li")
            for index, news1 in enumerate(news_list1) :
                a_img = 0
                img = news1.find("img")
                if img :
                    a_img = 1

                a_tag1 = news1.find_all("a")[a_img]
                title = a_tag1.get_text().strip()
                link = a_tag1["href"]
                press = news1.find("span", attrs={"class":"writing"}).get_text()

                page_index_1 = page_index_1 + 1

                ws1.append([page_index_1, title, link, press])

            page_index_1 = page_index_1 + index + 1

            if(soup.find("ul", attrs={"class":"type06"})) :
                news_list2 = soup.find("ul", attrs={"class":"type06"}).find_all("li")
                for index, news2 in enumerate(news_list2) :
                        a_img = 0
                        img = news2.find("img")
                        if img :
                            a_img = 1

                        a_tag2 = news2.find_all("a")[a_img]
                        title = a_tag2.get_text().strip()
                        link = a_tag2["href"]
                        press = news2.find("span", attrs={"class":"writing"}).get_text()

                        page_index_2 = page_index_2 + 1

                        ws1.append([page_index_2, title, link, press])

                page_index_2 = page_index_2 + index + 1

    page_number = 1
    paging_url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(page_number)
    paging_soup = create_paging_soup(paging_url)
    page_start = paging_soup.find("div", attrs={"class":"paging"}).find("strong").get_text()

    if(paging_soup.find("div", attrs={"class":"paging"}).find("a")) :
        page = paging_soup.find("div", attrs={"class":"paging"}).find_all("a")[-1]
        page_end = page.get_text()
    else :
        page_end = page_start

    if(page_end == "다음") :
        while(page_end == "다음") :
            page_number = page_number + 10
            paging_url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(page_number)
            paging_soup = create_paging_soup(paging_url)
            page = paging_soup.find("div", attrs={"class":"paging"}).find_all("a")[-1]
            page_end = page.get_text()

            if(page_end != "다음") :
                if(page_end == "이전") :
                    page_end = paging_soup.find("div", attrs={"class":"paging"}).find("strong").get_text()
                    scrape()
                else :
                    scrape()
    elif(page_end != "다음") :
        scrape()

def theme_103_245() :   # 생활/문화 > 생활문화일반
    sid1_code = str(103)
    sid2_code = str(245)

    ws1 = wb.create_sheet("생활,문화_생활문화일반")
    ws1 = wb["생활,문화_생활문화일반"]
    ws1.append(["index", "제목", "링크", "언론사"])

    def scrape() :
        page_index_1 = 0
        page_index_2 = 10

        for i in range(int(page_start), int(page_end)+1) :
            url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(i)

            soup = create_soup(url)
            news_list1 = soup.find("ul", attrs={"class":"type06_headline"}).find_all("li")
            for index, news1 in enumerate(news_list1) :
                a_img = 0
                img = news1.find("img")
                if img :
                    a_img = 1

                a_tag1 = news1.find_all("a")[a_img]
                title = a_tag1.get_text().strip()
                link = a_tag1["href"]
                press = news1.find("span", attrs={"class":"writing"}).get_text()

                page_index_1 = page_index_1 + 1

                ws1.append([page_index_1, title, link, press])

            page_index_1 = page_index_1 + index + 1

            if(soup.find("ul", attrs={"class":"type06"})) :
                news_list2 = soup.find("ul", attrs={"class":"type06"}).find_all("li")
                for index, news2 in enumerate(news_list2) :
                        a_img = 0
                        img = news2.find("img")
                        if img :
                            a_img = 1

                        a_tag2 = news2.find_all("a")[a_img]
                        title = a_tag2.get_text().strip()
                        link = a_tag2["href"]
                        press = news2.find("span", attrs={"class":"writing"}).get_text()

                        page_index_2 = page_index_2 + 1

                        ws1.append([page_index_2, title, link, press])

                page_index_2 = page_index_2 + index + 1

    page_number = 1
    paging_url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(page_number)
    paging_soup = create_paging_soup(paging_url)
    page_start = paging_soup.find("div", attrs={"class":"paging"}).find("strong").get_text()

    if(paging_soup.find("div", attrs={"class":"paging"}).find("a")) :
        page = paging_soup.find("div", attrs={"class":"paging"}).find_all("a")[-1]
        page_end = page.get_text()
    else :
        page_end = page_start

    if(page_end == "다음") :
        while(page_end == "다음") :
            page_number = page_number + 10
            paging_url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(page_number)
            paging_soup = create_paging_soup(paging_url)
            page = paging_soup.find("div", attrs={"class":"paging"}).find_all("a")[-1]
            page_end = page.get_text()

            if(page_end != "다음") :
                if(page_end == "이전") :
                    page_end = paging_soup.find("div", attrs={"class":"paging"}).find("strong").get_text()
                    scrape()
                else :
                    scrape()
    elif(page_end != "다음") :
        scrape()

def theme_104_231() :   # 세계 > 아시아/호주
    sid1_code = str(104)
    sid2_code = str(231)

    ws1 = wb.create_sheet("세계_아시아,호주")
    ws1 = wb["세계_아시아,호주"]
    ws1.append(["index", "제목", "링크", "언론사"])

    def scrape() :
        page_index_1 = 0
        page_index_2 = 10

        for i in range(int(page_start), int(page_end)+1) :
            url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(i)

            soup = create_soup(url)
            news_list1 = soup.find("ul", attrs={"class":"type06_headline"}).find_all("li")
            for index, news1 in enumerate(news_list1) :
                a_img = 0
                img = news1.find("img")
                if img :
                    a_img = 1

                a_tag1 = news1.find_all("a")[a_img]
                title = a_tag1.get_text().strip()
                link = a_tag1["href"]
                press = news1.find("span", attrs={"class":"writing"}).get_text()

                page_index_1 = page_index_1 + 1

                ws1.append([page_index_1, title, link, press])

            page_index_1 = page_index_1 + index + 1

            if(soup.find("ul", attrs={"class":"type06"})) :
                news_list2 = soup.find("ul", attrs={"class":"type06"}).find_all("li")
                for index, news2 in enumerate(news_list2) :
                        a_img = 0
                        img = news2.find("img")
                        if img :
                            a_img = 1

                        a_tag2 = news2.find_all("a")[a_img]
                        title = a_tag2.get_text().strip()
                        link = a_tag2["href"]
                        press = news2.find("span", attrs={"class":"writing"}).get_text()

                        page_index_2 = page_index_2 + 1

                        ws1.append([page_index_2, title, link, press])

                page_index_2 = page_index_2 + index + 1

    page_number = 1
    paging_url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(page_number)
    paging_soup = create_paging_soup(paging_url)
    page_start = paging_soup.find("div", attrs={"class":"paging"}).find("strong").get_text()

    if(paging_soup.find("div", attrs={"class":"paging"}).find("a")) :
        page = paging_soup.find("div", attrs={"class":"paging"}).find_all("a")[-1]
        page_end = page.get_text()
    else :
        page_end = page_start

    if(page_end == "다음") :
        while(page_end == "다음") :
            page_number = page_number + 10
            paging_url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(page_number)
            paging_soup = create_paging_soup(paging_url)
            page = paging_soup.find("div", attrs={"class":"paging"}).find_all("a")[-1]
            page_end = page.get_text()

            if(page_end != "다음") :
                if(page_end == "이전") :
                    page_end = paging_soup.find("div", attrs={"class":"paging"}).find("strong").get_text()
                    scrape()
                else :
                    scrape()
    elif(page_end != "다음") :
        scrape()

def theme_104_232() :   # 세계 > 미국/중남미
    sid1_code = str(104)
    sid2_code = str(232)

    ws1 = wb.create_sheet("세계_미국,중남미")
    ws1 = wb["세계_미국,중남미"]
    ws1.append(["index", "제목", "링크", "언론사"])

    def scrape() :
        page_index_1 = 0
        page_index_2 = 10

        for i in range(int(page_start), int(page_end)+1) :
            url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(i)

            soup = create_soup(url)
            news_list1 = soup.find("ul", attrs={"class":"type06_headline"}).find_all("li")
            for index, news1 in enumerate(news_list1) :
                a_img = 0
                img = news1.find("img")
                if img :
                    a_img = 1

                a_tag1 = news1.find_all("a")[a_img]
                title = a_tag1.get_text().strip()
                link = a_tag1["href"]
                press = news1.find("span", attrs={"class":"writing"}).get_text()

                page_index_1 = page_index_1 + 1

                ws1.append([page_index_1, title, link, press])

            page_index_1 = page_index_1 + index + 1

            if(soup.find("ul", attrs={"class":"type06"})) :
                news_list2 = soup.find("ul", attrs={"class":"type06"}).find_all("li")
                for index, news2 in enumerate(news_list2) :
                        a_img = 0
                        img = news2.find("img")
                        if img :
                            a_img = 1

                        a_tag2 = news2.find_all("a")[a_img]
                        title = a_tag2.get_text().strip()
                        link = a_tag2["href"]
                        press = news2.find("span", attrs={"class":"writing"}).get_text()

                        page_index_2 = page_index_2 + 1

                        ws1.append([page_index_2, title, link, press])

                page_index_2 = page_index_2 + index + 1

    page_number = 1
    paging_url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(page_number)
    paging_soup = create_paging_soup(paging_url)
    page_start = paging_soup.find("div", attrs={"class":"paging"}).find("strong").get_text()

    if(paging_soup.find("div", attrs={"class":"paging"}).find("a")) :
        page = paging_soup.find("div", attrs={"class":"paging"}).find_all("a")[-1]
        page_end = page.get_text()
    else :
        page_end = page_start

    if(page_end == "다음") :
        while(page_end == "다음") :
            page_number = page_number + 10
            paging_url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(page_number)
            paging_soup = create_paging_soup(paging_url)
            page = paging_soup.find("div", attrs={"class":"paging"}).find_all("a")[-1]
            page_end = page.get_text()

            if(page_end != "다음") :
                if(page_end == "이전") :
                    page_end = paging_soup.find("div", attrs={"class":"paging"}).find("strong").get_text()
                    scrape()
                else :
                    scrape()
    elif(page_end != "다음") :
        scrape()

def theme_104_233() :   # 세계 > 유럽
    sid1_code = str(104)
    sid2_code = str(233)

    ws1 = wb.create_sheet("세계_유럽")
    ws1 = wb["세계_유럽"]
    ws1.append(["index", "제목", "링크", "언론사"])

    def scrape() :
        page_index_1 = 0
        page_index_2 = 10

        for i in range(int(page_start), int(page_end)+1) :
            url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(i)

            soup = create_soup(url)
            news_list1 = soup.find("ul", attrs={"class":"type06_headline"}).find_all("li")
            for index, news1 in enumerate(news_list1) :
                a_img = 0
                img = news1.find("img")
                if img :
                    a_img = 1

                a_tag1 = news1.find_all("a")[a_img]
                title = a_tag1.get_text().strip()
                link = a_tag1["href"]
                press = news1.find("span", attrs={"class":"writing"}).get_text()

                page_index_1 = page_index_1 + 1

                ws1.append([page_index_1, title, link, press])

            page_index_1 = page_index_1 + index + 1

            if(soup.find("ul", attrs={"class":"type06"})) :
                news_list2 = soup.find("ul", attrs={"class":"type06"}).find_all("li")
                for index, news2 in enumerate(news_list2) :
                        a_img = 0
                        img = news2.find("img")
                        if img :
                            a_img = 1

                        a_tag2 = news2.find_all("a")[a_img]
                        title = a_tag2.get_text().strip()
                        link = a_tag2["href"]
                        press = news2.find("span", attrs={"class":"writing"}).get_text()

                        page_index_2 = page_index_2 + 1

                        ws1.append([page_index_2, title, link, press])

                page_index_2 = page_index_2 + index + 1

    page_number = 1
    paging_url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(page_number)
    paging_soup = create_paging_soup(paging_url)
    page_start = paging_soup.find("div", attrs={"class":"paging"}).find("strong").get_text()

    if(paging_soup.find("div", attrs={"class":"paging"}).find("a")) :
        page = paging_soup.find("div", attrs={"class":"paging"}).find_all("a")[-1]
        page_end = page.get_text()
    else :
        page_end = page_start

    if(page_end == "다음") :
        while(page_end == "다음") :
            page_number = page_number + 10
            paging_url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(page_number)
            paging_soup = create_paging_soup(paging_url)
            page = paging_soup.find("div", attrs={"class":"paging"}).find_all("a")[-1]
            page_end = page.get_text()

            if(page_end != "다음") :
                if(page_end == "이전") :
                    page_end = paging_soup.find("div", attrs={"class":"paging"}).find("strong").get_text()
                    scrape()
                else :
                    scrape()
    elif(page_end != "다음") :
        scrape()

def theme_104_234() :   # 세계 > 중동/아프리카
    sid1_code = str(104)
    sid2_code = str(234)

    ws1 = wb.create_sheet("세계_중동,아프리카")
    ws1 = wb["세계_중동,아프리카"]
    ws1.append(["index", "제목", "링크", "언론사"])

    def scrape() :
        page_index_1 = 0
        page_index_2 = 10

        for i in range(int(page_start), int(page_end)+1) :
            url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(i)

            soup = create_soup(url)
            news_list1 = soup.find("ul", attrs={"class":"type06_headline"}).find_all("li")
            for index, news1 in enumerate(news_list1) :
                a_img = 0
                img = news1.find("img")
                if img :
                    a_img = 1

                a_tag1 = news1.find_all("a")[a_img]
                title = a_tag1.get_text().strip()
                link = a_tag1["href"]
                press = news1.find("span", attrs={"class":"writing"}).get_text()

                page_index_1 = page_index_1 + 1

                ws1.append([page_index_1, title, link, press])

            page_index_1 = page_index_1 + index + 1

            if(soup.find("ul", attrs={"class":"type06"})) :
                news_list2 = soup.find("ul", attrs={"class":"type06"}).find_all("li")
                for index, news2 in enumerate(news_list2) :
                        a_img = 0
                        img = news2.find("img")
                        if img :
                            a_img = 1

                        a_tag2 = news2.find_all("a")[a_img]
                        title = a_tag2.get_text().strip()
                        link = a_tag2["href"]
                        press = news2.find("span", attrs={"class":"writing"}).get_text()

                        page_index_2 = page_index_2 + 1

                        ws1.append([page_index_2, title, link, press])

                page_index_2 = page_index_2 + index + 1

    page_number = 1
    paging_url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(page_number)
    paging_soup = create_paging_soup(paging_url)
    page_start = paging_soup.find("div", attrs={"class":"paging"}).find("strong").get_text()

    if(paging_soup.find("div", attrs={"class":"paging"}).find("a")) :
        page = paging_soup.find("div", attrs={"class":"paging"}).find_all("a")[-1]
        page_end = page.get_text()
    else :
        page_end = page_start

    if(page_end == "다음") :
        while(page_end == "다음") :
            page_number = page_number + 10
            paging_url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(page_number)
            paging_soup = create_paging_soup(paging_url)
            page = paging_soup.find("div", attrs={"class":"paging"}).find_all("a")[-1]
            page_end = page.get_text()

            if(page_end != "다음") :
                if(page_end == "이전") :
                    page_end = paging_soup.find("div", attrs={"class":"paging"}).find("strong").get_text()
                    scrape()
                else :
                    scrape()
    elif(page_end != "다음") :
        scrape()

def theme_104_322() :   # 세계 > 세계일반
    sid1_code = str(104)
    sid2_code = str(322)

    ws1 = wb.create_sheet("세계_세계일반")
    ws1 = wb["세계_세계일반"]
    ws1.append(["index", "제목", "링크", "언론사"])

    def scrape() :
        page_index_1 = 0
        page_index_2 = 10

        for i in range(int(page_start), int(page_end)+1) :
            url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(i)

            soup = create_soup(url)
            news_list1 = soup.find("ul", attrs={"class":"type06_headline"}).find_all("li")
            for index, news1 in enumerate(news_list1) :
                a_img = 0
                img = news1.find("img")
                if img :
                    a_img = 1

                a_tag1 = news1.find_all("a")[a_img]
                title = a_tag1.get_text().strip()
                link = a_tag1["href"]
                press = news1.find("span", attrs={"class":"writing"}).get_text()

                page_index_1 = page_index_1 + 1

                ws1.append([page_index_1, title, link, press])

            page_index_1 = page_index_1 + index + 1

            if(soup.find("ul", attrs={"class":"type06"})) :
                news_list2 = soup.find("ul", attrs={"class":"type06"}).find_all("li")
                for index, news2 in enumerate(news_list2) :
                        a_img = 0
                        img = news2.find("img")
                        if img :
                            a_img = 1

                        a_tag2 = news2.find_all("a")[a_img]
                        title = a_tag2.get_text().strip()
                        link = a_tag2["href"]
                        press = news2.find("span", attrs={"class":"writing"}).get_text()

                        page_index_2 = page_index_2 + 1

                        ws1.append([page_index_2, title, link, press])

                page_index_2 = page_index_2 + index + 1

    page_number = 1
    paging_url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(page_number)
    paging_soup = create_paging_soup(paging_url)
    page_start = paging_soup.find("div", attrs={"class":"paging"}).find("strong").get_text()

    if(paging_soup.find("div", attrs={"class":"paging"}).find("a")) :
        page = paging_soup.find("div", attrs={"class":"paging"}).find_all("a")[-1]
        page_end = page.get_text()
    else :
        page_end = page_start

    if(page_end == "다음") :
        while(page_end == "다음") :
            page_number = page_number + 10
            paging_url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(page_number)
            paging_soup = create_paging_soup(paging_url)
            page = paging_soup.find("div", attrs={"class":"paging"}).find_all("a")[-1]
            page_end = page.get_text()

            if(page_end != "다음") :
                if(page_end == "이전") :
                    page_end = paging_soup.find("div", attrs={"class":"paging"}).find("strong").get_text()
                    scrape()
                else :
                    scrape()
    elif(page_end != "다음") :
        scrape()

def theme_105_731() :   # IT/과학 > 모바일
    sid1_code = str(105)
    sid2_code = str(731)

    ws1 = wb.create_sheet("IT,과학_모바일")
    ws1 = wb["IT,과학_모바일"]
    ws1.append(["index", "제목", "링크", "언론사"])

    def scrape() :
        page_index_1 = 0
        page_index_2 = 10

        for i in range(int(page_start), int(page_end)+1) :
            url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(i)

            soup = create_soup(url)
            news_list1 = soup.find("ul", attrs={"class":"type06_headline"}).find_all("li")
            for index, news1 in enumerate(news_list1) :
                a_img = 0
                img = news1.find("img")
                if img :
                    a_img = 1

                a_tag1 = news1.find_all("a")[a_img]
                title = a_tag1.get_text().strip()
                link = a_tag1["href"]
                press = news1.find("span", attrs={"class":"writing"}).get_text()

                page_index_1 = page_index_1 + 1

                ws1.append([page_index_1, title, link, press])

            page_index_1 = page_index_1 + index + 1

            if(soup.find("ul", attrs={"class":"type06"})) :
                news_list2 = soup.find("ul", attrs={"class":"type06"}).find_all("li")
                for index, news2 in enumerate(news_list2) :
                        a_img = 0
                        img = news2.find("img")
                        if img :
                            a_img = 1

                        a_tag2 = news2.find_all("a")[a_img]
                        title = a_tag2.get_text().strip()
                        link = a_tag2["href"]
                        press = news2.find("span", attrs={"class":"writing"}).get_text()

                        page_index_2 = page_index_2 + 1

                        ws1.append([page_index_2, title, link, press])

                page_index_2 = page_index_2 + index + 1

    page_number = 1
    paging_url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(page_number)
    paging_soup = create_paging_soup(paging_url)
    page_start = paging_soup.find("div", attrs={"class":"paging"}).find("strong").get_text()

    if(paging_soup.find("div", attrs={"class":"paging"}).find("a")) :
        page = paging_soup.find("div", attrs={"class":"paging"}).find_all("a")[-1]
        page_end = page.get_text()
    else :
        page_end = page_start

    if(page_end == "다음") :
        while(page_end == "다음") :
            page_number = page_number + 10
            paging_url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(page_number)
            paging_soup = create_paging_soup(paging_url)
            page = paging_soup.find("div", attrs={"class":"paging"}).find_all("a")[-1]
            page_end = page.get_text()

            if(page_end != "다음") :
                if(page_end == "이전") :
                    page_end = paging_soup.find("div", attrs={"class":"paging"}).find("strong").get_text()
                    scrape()
                else :
                    scrape()
    elif(page_end != "다음") :
        scrape()

def theme_105_226() :   # IT/과학 > 인터넷/SNS
    sid1_code = str(105)
    sid2_code = str(226)

    ws1 = wb.create_sheet("IT,과학_인터넷,SNS")
    ws1 = wb["IT,과학_인터넷,SNS"]
    ws1.append(["index", "제목", "링크", "언론사"])

    def scrape() :
        page_index_1 = 0
        page_index_2 = 10

        for i in range(int(page_start), int(page_end)+1) :
            url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(i)

            soup = create_soup(url)
            news_list1 = soup.find("ul", attrs={"class":"type06_headline"}).find_all("li")
            for index, news1 in enumerate(news_list1) :
                a_img = 0
                img = news1.find("img")
                if img :
                    a_img = 1

                a_tag1 = news1.find_all("a")[a_img]
                title = a_tag1.get_text().strip()
                link = a_tag1["href"]
                press = news1.find("span", attrs={"class":"writing"}).get_text()

                page_index_1 = page_index_1 + 1

                ws1.append([page_index_1, title, link, press])

            page_index_1 = page_index_1 + index + 1

            if(soup.find("ul", attrs={"class":"type06"})) :
                news_list2 = soup.find("ul", attrs={"class":"type06"}).find_all("li")
                for index, news2 in enumerate(news_list2) :
                        a_img = 0
                        img = news2.find("img")
                        if img :
                            a_img = 1

                        a_tag2 = news2.find_all("a")[a_img]
                        title = a_tag2.get_text().strip()
                        link = a_tag2["href"]
                        press = news2.find("span", attrs={"class":"writing"}).get_text()

                        page_index_2 = page_index_2 + 1

                        ws1.append([page_index_2, title, link, press])

                page_index_2 = page_index_2 + index + 1

    page_number = 1
    paging_url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(page_number)
    paging_soup = create_paging_soup(paging_url)
    page_start = paging_soup.find("div", attrs={"class":"paging"}).find("strong").get_text()

    if(paging_soup.find("div", attrs={"class":"paging"}).find("a")) :
        page = paging_soup.find("div", attrs={"class":"paging"}).find_all("a")[-1]
        page_end = page.get_text()
    else :
        page_end = page_start

    if(page_end == "다음") :
        while(page_end == "다음") :
            page_number = page_number + 10
            paging_url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(page_number)
            paging_soup = create_paging_soup(paging_url)
            page = paging_soup.find("div", attrs={"class":"paging"}).find_all("a")[-1]
            page_end = page.get_text()

            if(page_end != "다음") :
                if(page_end == "이전") :
                    page_end = paging_soup.find("div", attrs={"class":"paging"}).find("strong").get_text()
                    scrape()
                else :
                    scrape()
    elif(page_end != "다음") :
        scrape()

def theme_105_227() :   # IT/과학 > 통신/뉴미디어
    sid1_code = str(105)
    sid2_code = str(227)

    ws1 = wb.create_sheet("IT,과학_통신,뉴미디어")
    ws1 = wb["IT,과학_통신,뉴미디어"]
    ws1.append(["index", "제목", "링크", "언론사"])

    def scrape() :
        page_index_1 = 0
        page_index_2 = 10

        for i in range(int(page_start), int(page_end)+1) :
            url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(i)

            soup = create_soup(url)
            news_list1 = soup.find("ul", attrs={"class":"type06_headline"}).find_all("li")
            for index, news1 in enumerate(news_list1) :
                a_img = 0
                img = news1.find("img")
                if img :
                    a_img = 1

                a_tag1 = news1.find_all("a")[a_img]
                title = a_tag1.get_text().strip()
                link = a_tag1["href"]
                press = news1.find("span", attrs={"class":"writing"}).get_text()

                page_index_1 = page_index_1 + 1

                ws1.append([page_index_1, title, link, press])

            page_index_1 = page_index_1 + index + 1

            if(soup.find("ul", attrs={"class":"type06"})) :
                news_list2 = soup.find("ul", attrs={"class":"type06"}).find_all("li")
                for index, news2 in enumerate(news_list2) :
                        a_img = 0
                        img = news2.find("img")
                        if img :
                            a_img = 1

                        a_tag2 = news2.find_all("a")[a_img]
                        title = a_tag2.get_text().strip()
                        link = a_tag2["href"]
                        press = news2.find("span", attrs={"class":"writing"}).get_text()

                        page_index_2 = page_index_2 + 1

                        ws1.append([page_index_2, title, link, press])

                page_index_2 = page_index_2 + index + 1

    page_number = 1
    paging_url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(page_number)
    paging_soup = create_paging_soup(paging_url)
    page_start = paging_soup.find("div", attrs={"class":"paging"}).find("strong").get_text()

    if(paging_soup.find("div", attrs={"class":"paging"}).find("a")) :
        page = paging_soup.find("div", attrs={"class":"paging"}).find_all("a")[-1]
        page_end = page.get_text()
    else :
        page_end = page_start

    if(page_end == "다음") :
        while(page_end == "다음") :
            page_number = page_number + 10
            paging_url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(page_number)
            paging_soup = create_paging_soup(paging_url)
            page = paging_soup.find("div", attrs={"class":"paging"}).find_all("a")[-1]
            page_end = page.get_text()

            if(page_end != "다음") :
                if(page_end == "이전") :
                    page_end = paging_soup.find("div", attrs={"class":"paging"}).find("strong").get_text()
                    scrape()
                else :
                    scrape()
    elif(page_end != "다음") :
        scrape()

def theme_105_230() :   # IT/과학 > IT일반
    sid1_code = str(105)
    sid2_code = str(230)

    ws1 = wb.create_sheet("IT,과학_IT일반")
    ws1 = wb["IT,과학_IT일반"]
    ws1.append(["index", "제목", "링크", "언론사"])

    def scrape() :
        page_index_1 = 0
        page_index_2 = 10

        for i in range(int(page_start), int(page_end)+1) :
            url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(i)

            soup = create_soup(url)
            news_list1 = soup.find("ul", attrs={"class":"type06_headline"}).find_all("li")
            for index, news1 in enumerate(news_list1) :
                a_img = 0
                img = news1.find("img")
                if img :
                    a_img = 1

                a_tag1 = news1.find_all("a")[a_img]
                title = a_tag1.get_text().strip()
                link = a_tag1["href"]
                press = news1.find("span", attrs={"class":"writing"}).get_text()

                page_index_1 = page_index_1 + 1

                ws1.append([page_index_1, title, link, press])

            page_index_1 = page_index_1 + index + 1

            if(soup.find("ul", attrs={"class":"type06"})) :
                news_list2 = soup.find("ul", attrs={"class":"type06"}).find_all("li")
                for index, news2 in enumerate(news_list2) :
                        a_img = 0
                        img = news2.find("img")
                        if img :
                            a_img = 1

                        a_tag2 = news2.find_all("a")[a_img]
                        title = a_tag2.get_text().strip()
                        link = a_tag2["href"]
                        press = news2.find("span", attrs={"class":"writing"}).get_text()

                        page_index_2 = page_index_2 + 1

                        ws1.append([page_index_2, title, link, press])

                page_index_2 = page_index_2 + index + 1

    page_number = 1
    paging_url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(page_number)
    paging_soup = create_paging_soup(paging_url)
    page_start = paging_soup.find("div", attrs={"class":"paging"}).find("strong").get_text()

    if(paging_soup.find("div", attrs={"class":"paging"}).find("a")) :
        page = paging_soup.find("div", attrs={"class":"paging"}).find_all("a")[-1]
        page_end = page.get_text()
    else :
        page_end = page_start

    if(page_end == "다음") :
        while(page_end == "다음") :
            page_number = page_number + 10
            paging_url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(page_number)
            paging_soup = create_paging_soup(paging_url)
            page = paging_soup.find("div", attrs={"class":"paging"}).find_all("a")[-1]
            page_end = page.get_text()

            if(page_end != "다음") :
                if(page_end == "이전") :
                    page_end = paging_soup.find("div", attrs={"class":"paging"}).find("strong").get_text()
                    scrape()
                else :
                    scrape()
    elif(page_end != "다음") :
        scrape()

def theme_105_732() :   # IT/과학 > 보안/해킹
    sid1_code = str(105)
    sid2_code = str(732)

    ws1 = wb.create_sheet("IT,과학_보안,해킹")
    ws1 = wb["IT,과학_보안,해킹"]
    ws1.append(["index", "제목", "링크", "언론사"])

    def scrape() :
        page_index_1 = 0
        page_index_2 = 10

        for i in range(int(page_start), int(page_end)+1) :
            url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(i)

            soup = create_soup(url)
            news_list1 = soup.find("ul", attrs={"class":"type06_headline"}).find_all("li")
            for index, news1 in enumerate(news_list1) :
                a_img = 0
                img = news1.find("img")
                if img :
                    a_img = 1

                a_tag1 = news1.find_all("a")[a_img]
                title = a_tag1.get_text().strip()
                link = a_tag1["href"]
                press = news1.find("span", attrs={"class":"writing"}).get_text()

                page_index_1 = page_index_1 + 1

                ws1.append([page_index_1, title, link, press])

            page_index_1 = page_index_1 + index + 1

            if(soup.find("ul", attrs={"class":"type06"})) :
                news_list2 = soup.find("ul", attrs={"class":"type06"}).find_all("li")
                for index, news2 in enumerate(news_list2) :
                        a_img = 0
                        img = news2.find("img")
                        if img :
                            a_img = 1

                        a_tag2 = news2.find_all("a")[a_img]
                        title = a_tag2.get_text().strip()
                        link = a_tag2["href"]
                        press = news2.find("span", attrs={"class":"writing"}).get_text()

                        page_index_2 = page_index_2 + 1

                        ws1.append([page_index_2, title, link, press])

                page_index_2 = page_index_2 + index + 1

    page_number = 1
    paging_url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(page_number)
    paging_soup = create_paging_soup(paging_url)
    page_start = paging_soup.find("div", attrs={"class":"paging"}).find("strong").get_text()

    if(paging_soup.find("div", attrs={"class":"paging"}).find("a")) :
        page = paging_soup.find("div", attrs={"class":"paging"}).find_all("a")[-1]
        page_end = page.get_text()
    else :
        page_end = page_start

    if(page_end == "다음") :
        while(page_end == "다음") :
            page_number = page_number + 10
            paging_url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(page_number)
            paging_soup = create_paging_soup(paging_url)
            page = paging_soup.find("div", attrs={"class":"paging"}).find_all("a")[-1]
            page_end = page.get_text()

            if(page_end != "다음") :
                if(page_end == "이전") :
                    page_end = paging_soup.find("div", attrs={"class":"paging"}).find("strong").get_text()
                    scrape()
                else :
                    scrape()
    elif(page_end != "다음") :
        scrape()

def theme_105_283() :   # IT/과학 > 컴퓨터
    sid1_code = str(105)
    sid2_code = str(283)

    ws1 = wb.create_sheet("IT,과학_컴퓨터")
    ws1 = wb["IT,과학_컴퓨터"]
    ws1.append(["index", "제목", "링크", "언론사"])

    def scrape() :
        page_index_1 = 0
        page_index_2 = 10

        for i in range(int(page_start), int(page_end)+1) :
            url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(i)

            soup = create_soup(url)
            news_list1 = soup.find("ul", attrs={"class":"type06_headline"}).find_all("li")
            for index, news1 in enumerate(news_list1) :
                a_img = 0
                img = news1.find("img")
                if img :
                    a_img = 1

                a_tag1 = news1.find_all("a")[a_img]
                title = a_tag1.get_text().strip()
                link = a_tag1["href"]
                press = news1.find("span", attrs={"class":"writing"}).get_text()

                page_index_1 = page_index_1 + 1

                ws1.append([page_index_1, title, link, press])

            page_index_1 = page_index_1 + index + 1

            if(soup.find("ul", attrs={"class":"type06"})) :
                news_list2 = soup.find("ul", attrs={"class":"type06"}).find_all("li")
                for index, news2 in enumerate(news_list2) :
                        a_img = 0
                        img = news2.find("img")
                        if img :
                            a_img = 1

                        a_tag2 = news2.find_all("a")[a_img]
                        title = a_tag2.get_text().strip()
                        link = a_tag2["href"]
                        press = news2.find("span", attrs={"class":"writing"}).get_text()

                        page_index_2 = page_index_2 + 1

                        ws1.append([page_index_2, title, link, press])

                page_index_2 = page_index_2 + index + 1

    page_number = 1
    paging_url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(page_number)
    paging_soup = create_paging_soup(paging_url)
    page_start = paging_soup.find("div", attrs={"class":"paging"}).find("strong").get_text()

    if(paging_soup.find("div", attrs={"class":"paging"}).find("a")) :
        page = paging_soup.find("div", attrs={"class":"paging"}).find_all("a")[-1]
        page_end = page.get_text()
    else :
        page_end = page_start

    if(page_end == "다음") :
        while(page_end == "다음") :
            page_number = page_number + 10
            paging_url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(page_number)
            paging_soup = create_paging_soup(paging_url)
            page = paging_soup.find("div", attrs={"class":"paging"}).find_all("a")[-1]
            page_end = page.get_text()

            if(page_end != "다음") :
                if(page_end == "이전") :
                    page_end = paging_soup.find("div", attrs={"class":"paging"}).find("strong").get_text()
                    scrape()
                else :
                    scrape()
    elif(page_end != "다음") :
        scrape()

def theme_105_229() :   # IT/과학 > 게임/리뷰
    sid1_code = str(105)
    sid2_code = str(229)

    ws1 = wb.create_sheet("IT,과학_게임,리뷰")
    ws1 = wb["IT,과학_게임,리뷰"]
    ws1.append(["index", "제목", "링크", "언론사"])

    def scrape() :
        page_index_1 = 0
        page_index_2 = 10

        for i in range(int(page_start), int(page_end)+1) :
            url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(i)

            soup = create_soup(url)
            news_list1 = soup.find("ul", attrs={"class":"type06_headline"}).find_all("li")
            for index, news1 in enumerate(news_list1) :
                a_img = 0
                img = news1.find("img")
                if img :
                    a_img = 1

                a_tag1 = news1.find_all("a")[a_img]
                title = a_tag1.get_text().strip()
                link = a_tag1["href"]
                press = news1.find("span", attrs={"class":"writing"}).get_text()

                page_index_1 = page_index_1 + 1

                ws1.append([page_index_1, title, link, press])

            page_index_1 = page_index_1 + index + 1

            if(soup.find("ul", attrs={"class":"type06"})) :
                news_list2 = soup.find("ul", attrs={"class":"type06"}).find_all("li")
                for index, news2 in enumerate(news_list2) :
                        a_img = 0
                        img = news2.find("img")
                        if img :
                            a_img = 1

                        a_tag2 = news2.find_all("a")[a_img]
                        title = a_tag2.get_text().strip()
                        link = a_tag2["href"]
                        press = news2.find("span", attrs={"class":"writing"}).get_text()

                        page_index_2 = page_index_2 + 1

                        ws1.append([page_index_2, title, link, press])

                page_index_2 = page_index_2 + index + 1

    page_number = 1
    paging_url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(page_number)
    paging_soup = create_paging_soup(paging_url)
    page_start = paging_soup.find("div", attrs={"class":"paging"}).find("strong").get_text()

    if(paging_soup.find("div", attrs={"class":"paging"}).find("a")) :
        page = paging_soup.find("div", attrs={"class":"paging"}).find_all("a")[-1]
        page_end = page.get_text()
    else :
        page_end = page_start

    if(page_end == "다음") :
        while(page_end == "다음") :
            page_number = page_number + 10
            paging_url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(page_number)
            paging_soup = create_paging_soup(paging_url)
            page = paging_soup.find("div", attrs={"class":"paging"}).find_all("a")[-1]
            page_end = page.get_text()

            if(page_end != "다음") :
                if(page_end == "이전") :
                    page_end = paging_soup.find("div", attrs={"class":"paging"}).find("strong").get_text()
                    scrape()
                else :
                    scrape()
    elif(page_end != "다음") :
        scrape()

def theme_105_228() :   # IT/과학 > 과학일반
    sid1_code = str(105)
    sid2_code = str(228)

    ws1 = wb.create_sheet("IT,과학_과학일반")
    ws1 = wb["IT,과학_과학일반"]
    ws1.append(["index", "제목", "링크", "언론사"])

    def scrape() :
        page_index_1 = 0
        page_index_2 = 10

        for i in range(int(page_start), int(page_end)+1) :
            url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(i)

            soup = create_soup(url)
            news_list1 = soup.find("ul", attrs={"class":"type06_headline"}).find_all("li")
            for index, news1 in enumerate(news_list1) :
                a_img = 0
                img = news1.find("img")
                if img :
                    a_img = 1

                a_tag1 = news1.find_all("a")[a_img]
                title = a_tag1.get_text().strip()
                link = a_tag1["href"]
                press = news1.find("span", attrs={"class":"writing"}).get_text()

                page_index_1 = page_index_1 + 1

                ws1.append([page_index_1, title, link, press])

            page_index_1 = page_index_1 + index + 1

            if(soup.find("ul", attrs={"class":"type06"})) :
                news_list2 = soup.find("ul", attrs={"class":"type06"}).find_all("li")
                for index, news2 in enumerate(news_list2) :
                        a_img = 0
                        img = news2.find("img")
                        if img :
                            a_img = 1

                        a_tag2 = news2.find_all("a")[a_img]
                        title = a_tag2.get_text().strip()
                        link = a_tag2["href"]
                        press = news2.find("span", attrs={"class":"writing"}).get_text()

                        page_index_2 = page_index_2 + 1

                        ws1.append([page_index_2, title, link, press])

                page_index_2 = page_index_2 + index + 1

    page_number = 1
    paging_url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(page_number)
    paging_soup = create_paging_soup(paging_url)
    page_start = paging_soup.find("div", attrs={"class":"paging"}).find("strong").get_text()

    if(paging_soup.find("div", attrs={"class":"paging"}).find("a")) :
        page = paging_soup.find("div", attrs={"class":"paging"}).find_all("a")[-1]
        page_end = page.get_text()
    else :
        page_end = page_start

    if(page_end == "다음") :
        while(page_end == "다음") :
            page_number = page_number + 10
            paging_url = "https://news.naver.com/main/list.naver?mode=LS2D&sid2=" + sid2_code + "&sid1=" + sid1_code + "&mid=shm&date=" + date + "&page=" + str(page_number)
            paging_soup = create_paging_soup(paging_url)
            page = paging_soup.find("div", attrs={"class":"paging"}).find_all("a")[-1]
            page_end = page.get_text()

            if(page_end != "다음") :
                if(page_end == "이전") :
                    page_end = paging_soup.find("div", attrs={"class":"paging"}).find("strong").get_text()
                    scrape()
                else :
                    scrape()
    elif(page_end != "다음") :
        scrape()

def scrape_news() :
    theme_100_264() # 정치_청와대
    theme_100_265() # 정치_국회/정당
    theme_100_268() # 정치_북한
    theme_100_266() # 정치_행정
    theme_100_267() # 정치_국방/외교
    theme_100_269() # 정치_정치일반

    theme_101_259() # 경제_금융
    theme_101_258() # 경제_증권
    theme_101_261() # 경제_산업/재계
    theme_101_771() # 경제_중기/벤처
    theme_101_260() # 경제_부동산
    theme_101_262() # 경제_글로벌경제
    theme_101_310() # 경제_생활경제
    theme_101_263() # 경제_경제일반

    theme_102_249() # 사회_사건사고
    theme_102_250() # 사회_교육
    theme_102_251() # 사회_노동
    theme_102_254() # 사회_언론
    theme_102_252() # 사회_환경
    theme_102_59b() # 사회_인권/복지
    theme_102_255() # 사회_식품/의료
    theme_102_256() # 사회_지역
    theme_102_276() # 사회_인물
    theme_102_257() # 사회_사회일반

    theme_103_241() # 생활/문화_건강정보
    theme_103_239() # 생활/문화_자동차/시승기
    theme_103_240() # 생활/문화_도로/교통
    theme_103_237() # 생활/문화_여행/레저
    theme_103_238() # 생활/문화_음식/맛집
    theme_103_376() # 생활/문화_패션/뷰티
    theme_103_242() # 생활/문화_공연/전시
    theme_103_243() # 생활/문화_책
    theme_103_244() # 생활/문화_종교
    theme_103_248() # 생활/문화_날씨
    theme_103_245() # 생활/문화_생활문화일반

    theme_104_231() # 세계_아시아/호주
    theme_104_232() # 세계_미국/중남미
    theme_104_233() # 세계_유럽
    theme_104_234() # 세계_중동/아프리카
    theme_104_322() # 세계_세계일반

    theme_105_731() # IT/과학_모바일
    theme_105_226() # IT/과학_인터넷/SNS
    theme_105_227() # IT/과학_통신/뉴미디어
    theme_105_230() # IT/과학_IT일반
    theme_105_732() # IT/과학_보안/해킹
    theme_105_283() # IT/과학_컴퓨터
    theme_105_229() # IT/과학_게임/리뷰
    theme_105_228() # IT/과학_과학일반

    wb.save("Crawiling {}.{}.{} {}시{}분{}초.xlsx".format(time.tm_year, time.tm_mon, time.tm_mday, time.tm_hour, time.tm_min, time.tm_sec))

if __name__ == "__main__" :
    scrape_news() # 뉴스 정보 가져오기
